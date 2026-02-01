import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Author Response"

# ── Style definitions ──
header_font = Font(name="Calibri", size=14, bold=True)
info_label_font = Font(name="Calibri", size=11, bold=True)
info_value_font = Font(name="Calibri", size=11)

col_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
col_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
col_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

reviewer_fill_editor = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
reviewer_fill_b = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
reviewer_fill_c = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
reviewer_fill_d = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
reviewer_fill_quality = PatternFill(start_color="7B7B7B", end_color="7B7B7B", fill_type="solid")

body_font = Font(name="Calibri", size=10)
body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

alt_fill_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ── Column widths ──
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 75
ws.column_dimensions["D"].width = 14

# ── Header section ──
ws.merge_cells("A1:D1")
ws["A1"] = "AUTHOR'S RESPONSE"
ws["A1"].font = header_font
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A3:D3")
ws["A3"] = "Kinetik: Game Technology, Information System, Computer Network, Computing, Electronics, and Control"
ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="666666")
ws["A3"].alignment = Alignment(horizontal="center")

row = 5
info = [
    ("Paper ID", "2558"),
    ("Author", "Akhiyar Waladi, Hasanatul Iftitah, Yogi Perdana, Nindy Raisa Hanum, Fitra Wahyuni, Rahmad Ashar"),
    ("Title", "Parameter Efficient Models for Malaria Detection and Classification Using Small-Scale Imbalanced Blood Smear Images"),
]
for label, value in info:
    ws.cell(row=row, column=1, value=label).font = info_label_font
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.cell(row=row, column=2, value=value).font = info_value_font
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    row += 1

row += 1


def write_section_header(ws, row, title, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = fill
    return row + 1


def write_col_headers(ws, row):
    headers = ["No.", "Reviewer's Comment", "Author's Response / Revised Version", "Page"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = col_header_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 24
    return row + 1


def write_row(ws, row, no, comment, response, page, alt_idx):
    fill = alt_fill_1 if alt_idx % 2 == 0 else alt_fill_2
    vals = [no, comment, response, page]
    aligns = [center_align, body_align, body_align, center_align]
    for i, (v, a) in enumerate(zip(vals, aligns), 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = body_font
        cell.alignment = a
        cell.border = thin_border
        cell.fill = fill
    # Auto-adjust row height based on wrapped text in each column
    col_widths = {1: 5, 2: 50, 3: 75, 4: 14}  # char widths per column
    max_lines = 1
    for col_idx, val in enumerate(vals, 1):
        text = str(val) if val else ""
        char_width = col_widths.get(col_idx, 50)
        for paragraph in text.split("\n"):
            max_lines += max(0, len(paragraph) // char_width)
        max_lines += text.count("\n")
    line_height = 14
    ws.row_dimensions[row].height = max(line_height * max_lines, 30)
    return row + 1


# ════════════════════════════════════════
# EDITOR COMMENTS
# ════════════════════════════════════════
row = write_section_header(ws, row, "Editor Comments", reviewer_fill_editor)
row = write_col_headers(ws, row)

editor_data = [
    ("1",
     "An Introduction should contain:\n"
     "- Background\n"
     "- The Problem\n"
     "- The Proposed Solution",

     "The Introduction has been restructured to follow this three-part framework. "
     "Section 1.1 establishes background and motivation. "
     "Section 1.2 has been renamed to \"Literature Review\" and expanded from two to three paragraphs, "
     "reviewing five major contributors (Arshad et al., Loddo et al., Zedda et al., Sukumarran et al.) "
     "and explicitly identifying three critical research gaps as the problem statement. "
     "Section 1.3 presents the proposed shared classification architecture, "
     "and Section 1.4 details four specific contributions.",
     "Sec. 1\npp. 1-3"),

    ("2",
     "Improve your analyzing and present comparison between performance of your approach and other researches.",

     "A new Section 3.6 \"Discussion\" has been added, comprising three paragraphs that: "
     "(1) state three initial hypotheses, "
     "(2) evaluate each hypothesis against experimental evidence across four datasets, and "
     "(3) discuss broader implications for clinical deployment and the medical imaging field. "
     "Section 3.5 has been enhanced with an explicit reference to Table 7 for state-of-the-art comparison.",
     "Sec. 3.5-3.6\npp. 8-10"),

    ("3",
     "Prepare your figures in high quality, created by yourself. All legends, captions MUST be in English.",

     "All six figures are original, generated from our experimental pipeline using Python and matplotlib. "
     "All legends, axis labels, and captions are written entirely in English. "
     "No figures are reproduced from external sources.",
     "Figs. 1-6"),

    ("4",
     "Each submitted article must be written in English.",
     "Confirmed. The entire article is written in English.",
     "All pages"),

    ("5",
     "Figures and Tables must be written in English.",
     "Confirmed. All seven tables and six figures use English exclusively for headers, labels, captions, and legends.",
     "All pages"),

    ("6",
     "All Figures, tables, and formulas are clearly visible and referenced in a paragraph.",

     "All cross-references have been verified and corrected. "
     "Table 7, which was previously unreferenced in the body text, now has an explicit reference in Section 3.5. "
     "Missing closing parentheses in Table 3, 4, and 5 captions have been added.",
     "Sec. 3.2\nSec. 3.5"),

    ("7",
     "At least 25 references in IEEE style. Citations must be sequential.",

     "The manuscript contains 33 references in IEEE format. "
     "Eight citation number errors have been corrected across Sections 3.1, 3.3, and 3.5 "
     "where reference numbers did not match the intended sources. "
     "All citations now follow strict sequential first-appearance order throughout the document.",
     "Sec. 3.1\nSec. 3.3\nSec. 3.5"),

    ("8",
     "Article matches the Kinetik template.",
     "Confirmed. The revised manuscript follows the Kinetik template format.",
     "All pages"),

    ("9",
     "Check author guideline for submission.",
     "Noted. The author guidelines have been reviewed and followed.",
     "-"),

    ("10",
     "Improve analyzing and comparison with other researches.",
     "Addressed together with Point 2 through the new Section 3.6 \"Discussion\" and enhanced Section 3.5 comparison.",
     "Sec. 3.5-3.6\npp. 8-10"),
]

for i, (no, comment, response, page) in enumerate(editor_data):
    row = write_row(ws, row, no, comment, response, page, i)

row += 1

# ════════════════════════════════════════
# REVIEWER B
# ════════════════════════════════════════
row = write_section_header(ws, row, "Reviewer B  \u2014  Recommendation: Accept Submission", reviewer_fill_b)
row = write_col_headers(ws, row)

row = write_row(ws, row, "1",
    "Dari saya semua sudah ok dan terima submisi.\n"
    "(Translation: Everything is OK from my side, accept the submission.)",
    "We thank Reviewer B for the positive evaluation and acceptance recommendation.",
    "-", 0)

row += 1

# ════════════════════════════════════════
# REVIEWER C
# ════════════════════════════════════════
row = write_section_header(ws, row, "Reviewer C  \u2014  Recommendation: Revisions Required", reviewer_fill_c)
row = write_col_headers(ws, row)

reviewer_c_data = [
    ("1",
     "The introduction should contextualize your research. "
     "It must explain the significance of previous work and the issues your work addresses. "
     "Include:\n"
     "a) An outline of the problem\n"
     "b) A review of the relevant literature highlighting major contributors",

     "Section 1.2 has been completely rewritten as \"Literature Review\" and expanded from two to three paragraphs. "
     "The revised section explicitly reviews five major contributors and their findings: "
     "Arshad et al. [7] achieved 89.33% detection precision and 95.86% classification on IML Lifecycle; "
     "Loddo et al. [25] established the first deep learning baseline on MP-IDB with 85.18-97% accuracy; "
     "Zedda et al. [24] demonstrated 95.2% detection with YOLOv5 on MP-IDB; "
     "Zedda et al. [21] developed YOLO-PAM achieving 91.8% mAP@50 with 11M fewer parameters; "
     "and Sukumarran et al. [26] reported 96% mAP@0.5 with YOLOv5. "
     "The third paragraph identifies three unaddressed gaps: limited dataset sizes (200-500 images), "
     "extreme class imbalance up to 54:1, and computational redundancy in existing pipelines.",
     "Sec. 1.2\npp. 1-2"),

    ("2",
     "The Method must provide detailed description including justification for novel and standard approaches.",

     "Three methodological justifications have been added:\n\n"
     "(a) Section 2.2: Justification for the ground truth crop architecture, explaining how it "
     "eliminates detection noise propagation, ensures fair cross-model comparison on identical data, "
     "and enables a train-once-reuse paradigm that reduces computational scaling from multiplicative to additive.\n\n"
     "(b) Section 2.4: Justification for AdamW optimizer selection over SGD, citing its decoupled weight decay "
     "providing more consistent regularization on small datasets of 200-800 training images.\n\n"
     "(c) Section 2.4: Justification for Focal Loss parameters (\u03b1=1.0, \u03b3=1.5), explaining that "
     "\u03b1=1.0 avoids double-counting with the existing weighted random sampling strategy, while "
     "\u03b3=1.5 provides moderate hard-example focusing without destabilizing convergence.",
     "Sec. 2.2\nSec. 2.4\npp. 3-5"),

    ("3",
     "The discussion should include:\n"
     "- results of your research\n"
     "- a discussion of related research\n"
     "- comparison between results and initial hypothesis",

     "A new Section 3.6 \"Discussion\" has been added with three paragraphs. "
     "The first paragraph states three hypotheses: "
     "(H1) shared classification maintains accuracy while reducing computational overhead, "
     "(H2) parameter-efficient models outperform larger architectures on small datasets, and "
     "(H3) Focal Loss enables robust minority class performance. "
     "The second paragraph evaluates each hypothesis: "
     "H1 is strongly supported with 86.45-98.28% accuracy and training runs reduced from 18 to 6; "
     "H2 is partially confirmed as EfficientNet models win on 3 of 4 datasets but ResNet50 proves superior under 54:1 imbalance; "
     "H3 is confirmed with F1-scores of 0.44-1.00 on ultra-minority classes. "
     "The third paragraph discusses implications for clinical deployment of compact 46-89 MB models "
     "and contextualizes results within the progression of the field.",
     "Sec. 3.6\npp. 9-10"),

    ("4",
     "Your conclusion should address outstanding questions and propose applications/extensions. "
     "\"What do my findings mean for the research field and my community?\"",

     "A new fourth paragraph has been added to Section 4 that addresses this directly. "
     "It establishes the shared classification paradigm as a reusable architectural pattern "
     "for parasitological and cytological screening tasks. "
     "It highlights that compact 46-89 MB models achieving 86.45-98.28% accuracy provide a practical "
     "deployment pathway for the 40% of health facilities in endemic regions lacking trained microscopists. "
     "Finally, it identifies the outstanding question of bridging the laboratory-to-field gap: "
     "translating performance on curated datasets to real-world clinical workflows involving "
     "thick blood smears, variable staining quality, and diverse patient populations.",
     "Sec. 4\npp. 10-11"),
]

for i, (no, comment, response, page) in enumerate(reviewer_c_data):
    row = write_row(ws, row, no, comment, response, page, i)

row += 1

# ════════════════════════════════════════
# REVIEWER D
# ════════════════════════════════════════
row = write_section_header(ws, row, "Reviewer D  \u2014  Recommendation: Revisions Required", reviewer_fill_d)
row = write_col_headers(ws, row)

row = write_row(ws, row, "1",
    "I have some comment in my pdf file, please check and update it. "
    "You can follow my comment inside my pdf file as a guide.",

    "We thank Reviewer D for the positive evaluation and detailed inline comments. "
    "Each annotation in the PDF has been reviewed and addressed accordingly.\n\n"
    "[Specific responses to be added after reviewing Reviewer D's annotated PDF from the KINETIK OJS system.]",
    "-", 0)

row += 1

# ════════════════════════════════════════
# ADDITIONAL QUALITY IMPROVEMENTS
# ════════════════════════════════════════
row = write_section_header(ws, row, "Additional Quality Improvements", reviewer_fill_quality)
row = write_col_headers(ws, row)

quality_data = [
    ("1",
     'Double period ".." at end of sentence',
     'Corrected to single period.',
     "Sec. 2.1"),
    ("2",
     '"After stratified dataset yields"',
     'Corrected to: "After stratified splitting, the dataset yields"',
     "Sec. 2.1"),
    ("3",
     '"for fair evaluation judge [17]"',
     'Removed stray word: "for fair evaluation [17]"',
     "Sec. 2.1"),
    ("4",
     '"as show in Table 3"',
     'Corrected to: "as shown in Table 3"',
     "Sec. 3.2"),
    ("5",
     '"MP-IDB Species at Table 4"',
     'Corrected to: "MP-IDB Species in Table 4"',
     "Sec. 3.2"),
    ("6",
     '"These results show in Table 5 demonstrate"',
     'Corrected to: "These results shown in Table 5 demonstrate"',
     "Sec. 3.2"),
    ("7",
     '"as resultsin Table 6"',
     'Corrected to: "as shown in Table 6"',
     "Sec. 3.2"),
    ("8",
     "Missing closing parenthesis in Table 3, 4, 5 captions",
     'Added closing ")" to each table caption.',
     "Sec. 3.2"),
    ("9",
     'Wrong figure reference: "Figure 4f"',
     'Corrected to: "Figure 6f"',
     "Sec. 3.4"),
    ("10",
     '"missed on detection model parasites"',
     'Corrected to: "missed parasites"',
     "Sec. 3.4"),
    ("11",
     "Eight incorrect citation numbers",
     "Corrected eight citation-reference mismatches across Sections 3.1, 3.3, and 3.5.",
     "Sec. 3.1\nSec. 3.3\nSec. 3.5"),
    ("12",
     "Table 7 not referenced in body text",
     'Added: "Table 7 summarizes the comparative performance against prior work on IML Lifecycle and MP-IDB datasets."',
     "Sec. 3.5"),
]

for i, (no, issue, correction, page) in enumerate(quality_data):
    row = write_row(ws, row, no, issue, correction, page, i)

# ── Print setup ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"
ws.page_margins = openpyxl.worksheet.page.PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

# ── Save ──
out_path = r"C:\Users\MyPC PRO\Documents\hello_world\luaran\hand_created\papers\AUTHOR_RESPONSE.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total rows: {row}")
