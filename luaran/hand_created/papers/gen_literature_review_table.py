#!/usr/bin/env python3
"""
Generate Literature Review Summary Tables for KINETIK Paper (Section 1.2)

Produces two formatted Excel sheets:
  Sheet 1 - "Methodology-Gap" : Focus on methods, contributions, and limitations
  Sheet 2 - "RQ Alignment"    : Maps prior work to the three research gaps

Output: luaran/hand_created/papers/exports/Literature_Review_Table.xlsx

Usage:
    python luaran/hand_created/papers/gen_literature_review_table.py
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Paths ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(SCRIPT_DIR, "exports")
OUTPUT_FILE = os.path.join(EXPORTS_DIR, "Literature_Review_Table.xlsx")

os.makedirs(EXPORTS_DIR, exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# DATA — verified against original papers and PDF references (page 210)
# ══════════════════════════════════════════════════════════════════════════════

STUDIES = [
    {
        "ref": "[5]",
        "authors": "Arshad et al.",
        "year": 2022,
        "title": "A dataset and benchmark for malaria life-cycle classification in thin blood smear images",
        "venue": "Neural Comput. Appl.",
        # Combined method
        "method": "Det: Morphological segmentation + watershed\nCls: ResNet50V2",
        # Dataset
        "dataset": "IML (345 imgs, 4 P. vivax stages)",
        # Contribution & limitation (concise)
        "contribution": "Introduced IML dataset with bbox annotations; two-stage det-cls pipeline",
        "limitation": "Single species; small dataset; no cross-dataset eval; no imbalance handling",
        # RQ mapping
        "rq1_cross_dataset": "Single dataset\n(IML only)",
        "rq2_imbalance": "Not addressed\n(standard CE loss)",
        "rq3_efficiency": "Per-detector training\n(no shared architecture)",
    },
    {
        "ref": "[6]",
        "authors": "Loddo et al.",
        "year": 2022,
        "title": "An Empirical Evaluation of Convolutional Networks for Malaria Diagnosis",
        "venue": "J. Imaging",
        "method": "Cls only: 11 CNNs evaluated\n(best: DenseNet-201, Acc 99.40%)",
        "dataset": "MP-IDB + NIH (209 + 27,558 imgs)",
        "contribution": "First DL baseline on MP-IDB; comprehensive 11 CNN comparison",
        "limitation": "No detection component; single species (P. falciparum); no end-to-end workflow",
        "rq1_cross_dataset": "Single dataset\n(MP-IDB only)",
        "rq2_imbalance": "Not addressed",
        "rq3_efficiency": "Classification only\n(no detection pipeline)",
    },
    {
        "ref": "[7]",
        "authors": "Zedda et al.",
        "year": 2022,
        "title": "A Deep Learning Based Framework for Malaria Diagnosis on High Variation Data Set",
        "venue": "ICIAP 2022, Springer",
        "method": "Det: YOLOv5\nCls: DarkNet-53 (Acc 96.02%)",
        "dataset": "MP-IDB (209 imgs, P. falciparum)",
        "contribution": "First YOLO detection on MP-IDB; single-stage det matches cls-only pipelines",
        "limitation": "Single dataset; no imbalance handling; separate det and cls models",
        "rq1_cross_dataset": "Single dataset\n(MP-IDB only)",
        "rq2_imbalance": "Not addressed",
        "rq3_efficiency": "Per-detector training\n(separate cls per det)",
    },
    {
        "ref": "[8]",
        "authors": "Zedda et al.",
        "year": 2023,
        "title": "YOLO-PAM: Parasite-Attention-Based Model for Efficient Malaria Parasite Detection",
        "venue": "J. Imaging",
        "method": "Det+Cls: YOLO-PAM\n(YOLOv8 + NAM/CBAM attention)\nmAP@50: 91.8% IML, 83.6% MP-IDB",
        "dataset": "IML + MP-IDB (522 imgs, 2 species)",
        "contribution": "Attention mechanisms for malaria det; 11M fewer params vs YOLOv8",
        "limitation": "Detection-only eval; specialized arch limits reproducibility; no imbalance handling",
        "rq1_cross_dataset": "2 datasets\n(IML + MP-IDB)",
        "rq2_imbalance": "Not addressed",
        "rq3_efficiency": "Parameter reduction\n(attention pruning)\nbut no shared cls arch",
    },
    {
        "ref": "[9]",
        "authors": "Sukumarran et al.",
        "year": 2024,
        "title": "An optimised YOLOv4 deep learning model for efficient malarial cell detection",
        "venue": "Parasit. Vectors",
        "method": "Det: YOLOv4/v5 (mAP@50: 96%)\nCls: DenseNet-121 (Acc 95.5%)",
        "dataset": "IML + MP-IDB (522 imgs, 2 species)",
        "contribution": "YOLOv4 vs v5 comparison; two-stage det then species cls",
        "limitation": "Fixed cls arch (no multi-model comparison); no imbalance handling; per-detector training",
        "rq1_cross_dataset": "2 datasets\n(IML + MP-IDB)",
        "rq2_imbalance": "Not addressed",
        "rq3_efficiency": "Per-detector training\n(no shared architecture)",
    },
    {
        "ref": "Ours",
        "authors": "This study",
        "year": 2025,
        "title": "Parameter Efficient Models for Malaria Detection and Classification "
                 "Using Small-Scale Imbalanced Blood Smear Images",
        "venue": "KINETIK",
        "method": "Det: YOLOv10/v11/v12 (20.1M params)\nCls: 6 CNNs (DenseNet, EfficientNet, ResNet)",
        "dataset": "4 datasets, 1,544 imgs total\n(IML, MP-IDB Species/Stages, MD-2019)",
        "contribution": "Shared cls arch (train-once-reuse); 3x6 multi-model eval; "
                        "Focal Loss for imbalance (up to 54:1); 46-89 MB models",
        "limitation": "4 public datasets (max 813 imgs); bbox only; lab images only",
        "rq1_cross_dataset": "4 complementary datasets\n(1,544 total images;\nlifecycle + species + multi-patient)",
        "rq2_imbalance": "Focal Loss (a=1.0, g=1.5)\n+ weighted random sampling\nF1: 0.44-1.00 on minorities",
        "rq3_efficiency": "Shared classification\n(train-once-reuse paradigm)\n"
                          "EfficientNet 5.3-9.2M params\n46-89 MB model size",
    },
]


# ══════════════════════════════════════════════════════════════════════════════
# STYLE DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════

# Colors
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
OURS_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ALT_FILL_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ALT_FILL_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GAP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow for gaps
ADDRESSED_FILL = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")  # light green

# Fonts
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=9)
OURS_FONT = Font(name="Calibri", size=9, bold=True)
TITLE_FONT = Font(name="Calibri", size=12, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="666666")

# Alignment
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

# Border
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)


def apply_cell_style(cell, font=BODY_FONT, align=BODY_ALIGN, fill=None, border=THIN_BORDER):
    """Apply style to a cell."""
    cell.font = font
    cell.alignment = align
    cell.border = border
    if fill:
        cell.fill = fill


def auto_row_height(ws, row, col_widths, font_size=9):
    """Estimate and set row height based on text length and column width."""
    max_lines = 1
    for col_idx, width in enumerate(col_widths, 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            text = str(cell.value)
            # Count explicit newlines
            explicit_lines = text.count("\n") + 1
            # Estimate wrapped lines per segment
            chars_per_line = max(int(width * 1.3), 10)
            total_lines = 0
            for segment in text.split("\n"):
                total_lines += max(1, -(-len(segment) // chars_per_line))  # ceil division
            max_lines = max(max_lines, total_lines)
    ws.row_dimensions[row].height = max(15, max_lines * (font_size + 5))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Methodology-Gap Matrix
# ══════════════════════════════════════════════════════════════════════════════

def create_methodology_gap_sheet(wb):
    """Create Sheet 1: Methodology-Gap analysis table."""
    ws = wb.active
    ws.title = "Methodology-Gap"

    # Column definitions
    columns = [
        ("Study (Year)", 18),
        ("Method", 30),
        ("Dataset(s) & Scale", 22),
        ("Key Contribution", 34),
        ("Limitation / Gap", 34),
    ]

    col_widths = [w for _, w in columns]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1,
                         value="Table X  Summary of Related Work: Methodology and Identified Gaps")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1,
                       value="Bold row = this study. Yellow-highlighted cells in Limitation column "
                             "indicate research gaps addressed by the proposed framework.")
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row (row 3)
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        apply_cell_style(cell, font=HEADER_FONT, align=HEADER_ALIGN, fill=HEADER_FILL)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 32

    # Data rows
    data_start = header_row + 1
    for i, study in enumerate(STUDIES):
        row = data_start + i
        is_ours = study["ref"] == "Ours"
        row_fill = OURS_FILL if is_ours else (ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2)
        row_font = OURS_FONT if is_ours else BODY_FONT

        values = [
            f"{study['authors']} ({study['year']})\n{study['ref']}",
            study["method"],
            study["dataset"],
            study["contribution"],
            study["limitation"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            apply_cell_style(cell, font=row_font, align=BODY_ALIGN, fill=row_fill)

        auto_row_height(ws, row, col_widths)

    # Freeze panes
    ws.freeze_panes = "A4"

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Research Question Alignment
# ══════════════════════════════════════════════════════════════════════════════

def create_rq_alignment_sheet(wb):
    """Create Sheet 2: RQ alignment table mapping studies to three gaps."""
    ws = wb.create_sheet("RQ Alignment")

    columns = [
        ("Study (Year)", 18),
        ("Key Method", 24),
        ("Gap 1:\nCross-Dataset\nRobustness", 24),
        ("Gap 2:\nClass Imbalance\nHandling", 24),
        ("Gap 3:\nComputational\nEfficiency", 24),
    ]

    col_widths = [w for _, w in columns]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1,
                         value="Table X  Alignment of Prior Work with Identified Research Gaps")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1,
                       value="Green cells = gap addressed. Yellow cells = gap not addressed. "
                             "Bold row = this study.")
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    header_row = 3
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        apply_cell_style(cell, font=HEADER_FONT, align=HEADER_ALIGN, fill=HEADER_FILL)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 40

    # Data rows
    data_start = header_row + 1
    for i, study in enumerate(STUDIES):
        row = data_start + i
        is_ours = study["ref"] == "Ours"
        base_fill = OURS_FILL if is_ours else (ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2)
        row_font = OURS_FONT if is_ours else BODY_FONT

        # Use first line of method as key method summary
        key_method = study["method"].split("\n")[0]

        rq_values = [
            study["rq1_cross_dataset"],
            study["rq2_imbalance"],
            study["rq3_efficiency"],
        ]

        # Fixed columns
        fixed_values = [
            f"{study['authors']} ({study['year']})\n{study['ref']}",
            key_method,
        ]

        for col_idx, val in enumerate(fixed_values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            apply_cell_style(cell, font=row_font, align=BODY_ALIGN, fill=base_fill)

        # RQ columns with conditional coloring
        for rq_idx, rq_val in enumerate(rq_values):
            col_idx = 3 + rq_idx
            cell = ws.cell(row=row, column=col_idx, value=rq_val)

            if is_ours:
                cell_fill = ADDRESSED_FILL
            elif "Not addressed" in rq_val or "not addressed" in rq_val.lower():
                cell_fill = GAP_FILL
            elif "Single dataset" in rq_val or "Per-detector" in rq_val or "Classification only" in rq_val:
                cell_fill = GAP_FILL
            else:
                cell_fill = base_fill

            apply_cell_style(cell, font=row_font, align=BODY_ALIGN, fill=cell_fill)

        auto_row_height(ws, row, col_widths)

    # Freeze panes
    ws.freeze_panes = "A4"

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()

    print("Generating literature review tables...")
    print(f"  Studies: {len(STUDIES)} (5 prior works + this study)")

    create_methodology_gap_sheet(wb)
    print("  Sheet 1: Methodology-Gap Matrix ... done")

    create_rq_alignment_sheet(wb)
    print("  Sheet 2: RQ Alignment ... done")

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"  File size: {os.path.getsize(OUTPUT_FILE) / 1024:.1f} KB")
    print("\nSheets:")
    print("  1. Methodology-Gap  — 5 columns: Study (Year), Method, Dataset,")
    print("                        Key Contribution, Limitation/Gap")
    print("  2. RQ Alignment     — 5 columns: Study (Year), Key Method,")
    print("                        Gap1 (Cross-Dataset), Gap2 (Imbalance), Gap3 (Efficiency)")


if __name__ == "__main__":
    main()
