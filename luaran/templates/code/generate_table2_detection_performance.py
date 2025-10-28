"""
Generate Table 2: YOLO Detection Performance Comparison
Extracts detection results from experiment folders and creates formatted Excel table
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import json

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_font = Font(bold=True, color="FFFFFF", size=10)
data_font = Font(size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Detection data from experiments
# VERIFIED against: results/optA_20251016_200330/consolidated_analysis/cross_dataset_comparison/detection_performance_all_datasets.csv
detection_results = {
    'IML Lifecycle (4 stages)': {
        'YOLOv10': {'mAP@50': 93.81, 'mAP@50-95': 77.71, 'Precision': 90.22, 'Recall': 92.22},
        'YOLOv11': {'mAP@50': 94.99, 'mAP@50-95': 77.76, 'Precision': 91.91, 'Recall': 91.11},
        'YOLOv12': {'mAP@50': 94.40, 'mAP@50-95': 78.21, 'Precision': 92.20, 'Recall': 86.67}
    },
    'MP-IDB Species (4 species)': {
        'YOLOv10': {'mAP@50': 92.44, 'mAP@50-95': 60.12, 'Precision': 85.67, 'Recall': 90.73},
        'YOLOv11': {'mAP@50': 92.57, 'mAP@50-95': 62.17, 'Precision': 86.52, 'Recall': 91.88},
        'YOLOv12': {'mAP@50': 92.72, 'mAP@50-95': 62.25, 'Precision': 88.99, 'Recall': 89.28}
    },
    'MP-IDB Stages (4 stages)': {
        'YOLOv10': {'mAP@50': 93.78, 'mAP@50-95': 44.48, 'Precision': 91.57, 'Recall': 93.12},
        'YOLOv11': {'mAP@50': 94.48, 'mAP@50-95': 60.34, 'Precision': 93.15, 'Recall': 88.36},
        'YOLOv12': {'mAP@50': 96.27, 'mAP@50-95': 61.53, 'Precision': 92.91, 'Recall': 92.59}
    },
    'MD_2019 Stages (3 stages)': {
        'YOLOv10': {'mAP@50': 70.84, 'mAP@50-95': 57.05, 'Precision': 67.89, 'Recall': 71.05},
        'YOLOv11': {'mAP@50': 72.91, 'mAP@50-95': 57.71, 'Precision': 68.58, 'Recall': 75.70},
        'YOLOv12': {'mAP@50': 71.12, 'mAP@50-95': 56.93, 'Precision': 65.92, 'Recall': 75.18}
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Detection Performance"

# Row 1: Main header (Dataset + 3 YOLO models) - LEVEL 1
row_idx = 1
ws.cell(row_idx, 1).value = 'Dataset'
ws.cell(row_idx, 1).fill = header_fill
ws.cell(row_idx, 1).font = header_font
ws.cell(row_idx, 1).alignment = center_align
ws.cell(row_idx, 1).border = border

col_idx = 2
for model in ['YOLOv10', 'YOLOv11', 'YOLOv12']:
    # Merge 4 columns for each model (mAP@50, mAP@50-95, Precision, Recall)
    ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx+3)
    cell = ws.cell(row_idx, col_idx)
    cell.value = f'{model}\nMedium (20.1M params)'
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border
    col_idx += 4

# Row 2: Metric subheaders - LEVEL 2
row_idx = 2
ws.cell(row_idx, 1).value = ''
ws.cell(row_idx, 1).border = border

col_idx = 2
for model in ['YOLOv10', 'YOLOv11', 'YOLOv12']:
    for metric in ['mAP@50\n(%)', 'mAP@50-95\n(%)', 'Precision\n(%)', 'Recall\n(%)']:
        cell = ws.cell(row_idx, col_idx)
        cell.value = metric
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border
        col_idx += 1

# Data rows (one row per dataset)
row_idx = 3
for dataset_name, models_data in detection_results.items():
    # Dataset name in first column
    cell = ws.cell(row_idx, 1)
    cell.value = dataset_name
    cell.font = data_font
    cell.alignment = left_align
    cell.border = border

    # Metrics for each YOLO model
    col_idx = 2
    for model in ['YOLOv10', 'YOLOv11', 'YOLOv12']:
        metrics = models_data[model]
        for metric_key in ['mAP@50', 'mAP@50-95', 'Precision', 'Recall']:
            cell = ws.cell(row_idx, col_idx)
            cell.value = metrics[metric_key]
            cell.font = data_font
            cell.alignment = center_align
            cell.border = border
            # Format as number with 2 decimals
            cell.number_format = '0.00'
            col_idx += 1

    row_idx += 1

# Set column widths
ws.column_dimensions['A'].width = 26  # Dataset names
for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
    ws.column_dimensions[col_letter].width = 11  # Metric columns

# Set row heights
ws.row_dimensions[1].height = 35  # Main header
ws.row_dimensions[2].height = 30  # Metric subheaders
for row in range(3, 7):  # Data rows
    ws.row_dimensions[row].height = 25

# Save workbook
output_path = '../tables/Table2_Detection_Performance.xlsx'
wb.save(output_path)

print("✅ Table 2: YOLO Detection Performance created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Table Structure (2-Level Header):")
print("  - Format: Datasets in ROWS, YOLO models in COLUMNS")
print("  - 4 datasets × 3 YOLO models × 4 metrics = 4 rows × 13 columns")
print("  - Level 1 (Row 1): Model headers (YOLOv10, YOLOv11, YOLOv12)")
print("  - Level 2 (Row 2): Metric subheaders (mAP@50, mAP@50-95, Precision, Recall)")
print("  - Rows 3-6: Dataset results")
print()
print("💡 Data Source:")
print("  - Extracted from experiment detection results")
print("  - YOLOv10/v11/v12 Medium (20.1M parameters)")
print("  - Trained for 100 epochs on 640×640 images")
print()
print("📌 Key Highlights:")
print("  - IML Lifecycle: Best mAP@50 = 94.99% (YOLO11)")
print("  - MP-IDB Species: Best mAP@50 = 92.72% (YOLO12)")
print("  - MP-IDB Stages: Best mAP@50 = 96.27% (YOLO12) ⭐ BEST OVERALL")
print("  - MD_2019: Best mAP@50 = 72.91% (YOLO11)")
print()
print("✨ Format similar to Table 3 (Classification) - 2-level headers")
