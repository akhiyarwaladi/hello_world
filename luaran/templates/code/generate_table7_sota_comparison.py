"""
Create Table 6 - FAIR COMPARISON with 5 Papers (ALL use SAME datasets!)
Only papers using IML or MP-IDB datasets
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
our_work_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
our_work_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "SOTA Comparison"

# Headers (6 columns - method + metrics merged)
headers = ['References', 'Year', 'Dataset Used\n(Same as Ours!)',
           'Detection\n(Method + mAP@50%)',
           'Classification\n(Method + Accuracy%)',
           'Key Features']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows - 5 papers ONLY using SAME datasets
data = [
    ['Arshad et al. [30]',
     '2022',
     '✅ IML Lifecycle\n(313 images)',
     'Morphological\nsegmentation\nPrecision: 89.33%',
     'ResNet50V2\nLifecycle classification\nAccuracy: 95.86%',
     'Two-stage: segmentation then classification. IML dataset from Pakistan (38K cells). P. vivax only.'],

    ['Zedda et al. [34]',
     '2022',
     '✅ MP-IDB\n(209 images)',
     'YOLOv5\n(detection + classification)\nDetection mAP: N/R',
     'YOLOv5: 95.2%\nDarkNet-53: 96.02%\n(P. falciparum 4 stages)',
     'Real-time detector + classifier. Four-class lifecycle stages (ring, trophozoite, schizont, gametocyte).'],

    ['Loddo et al. [31]',
     '2022',
     '✅ MP-IDB\n(209 images)',
     'N/R',
     'VGG-19: 85.18% (binary)\nDenseNet-201: >85%\n(4 lifecycle stages)',
     'Binary + multi-class classification. First baseline for lifecycle stages on MP-IDB. P. falciparum focus.'],

    ['Zedda et al. [32]',
     '2023',
     '✅ IML: 313\n✅ MP-IDB: 209',
     'YOLO-PAM\n(YOLOv8 + NAM/CBAM)\nmAP@50: 91.8% (IML)\nmAP: 83.6% (MP-IDB)',
     'Not specified\n(detection only)',
     'Attention mechanisms (NAM/CBAM). Multi-dataset. Parameter-efficient (11M fewer params vs baseline).'],

    ['Sukumarran et al. [35]',
     '2024',
     '✅ IML: 313\n✅ MP-IDB: 209',
     'YOLOv5: 96% mAP@0.5\nYOLOv4: 89-90% mAP@0.5\n(source + validation)',
     'DenseNet-121\nSpecies identification\nAccuracy: 95.5%',
     'Two-stage: detection then species classification. Superior generalization (YOLOv4 on validation set).'],

    # Our Work - LAST ROW
    ['This Study',
     '2025',
     '✅ IML: 313\n✅ MP-IDB Species: 209\n✅ MP-IDB Stages: 209\n✅ MD_2019: 883\nTotal: 1,614',
     'YOLOv11 Medium\n(shared across datasets)\nmAP@50: 92.57-94.99%\nP: 68.58-92.91%\nR: 75.70-92.59%',
     'EfficientNet-B0/B1\nResNet50\nFocal Loss (α=0.25, γ=2.0)\nAcc: 84.22-98.28%\nBal.Acc: 83.04-91.96%\nF1 (minorities): ≥0.80',
     'Shared architecture (67% efficiency). Multi-dataset (4 datasets). Focal Loss for extreme imbalance (54:1). Per-class metrics. First MD_2019 with deep learning.']
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border

        # Style for Our Work (last row)
        if row_idx == len(data) + 1:  # Last row
            cell.fill = our_work_fill
            cell.font = our_work_font
            if col_idx == 2:  # Year column
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        else:
            if col_idx == 2:  # Year column
                cell.alignment = center_align
            else:
                cell.alignment = left_align

# Set column widths
ws.column_dimensions['A'].width = 22  # References
ws.column_dimensions['B'].width = 8   # Year
ws.column_dimensions['C'].width = 24  # Dataset (emphasized!)
ws.column_dimensions['D'].width = 28  # Detection (method + metrics)
ws.column_dimensions['E'].width = 32  # Classification (method + metrics)
ws.column_dimensions['F'].width = 48  # Key Features

# Set row heights for readability
ws.row_dimensions[1].height = 40  # Header
for row in range(2, len(data) + 2):
    if row == len(data) + 1:  # Our Work row (last)
        ws.row_dimensions[row].height = 100  # Taller for Our Work
    else:
        ws.row_dimensions[row].height = 75  # Other rows

# Save workbook
output_path = '../tables/Table7_Comparison_SOTA.xlsx'
wb.save(output_path)

print("✅ Table 7: FAIR COMPARISON (5 Papers - ALL Same Datasets!) created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Structure:")
print("  - 5 comparison studies (ALL use IML and/or MP-IDB!)")
print("  - Our Work at bottom (highlighted in yellow)")
print("  - 6 columns (method + metrics merged)")
print()
print("✅ FAIR COMPARISON:")
print("  1. Arshad 2022 - ✅ IML Dataset (313)")
print("  2. Zedda 2022 (ICIAP) - ✅ MP-IDB Dataset (209)")
print("  3. Loddo 2022 - ✅ MP-IDB Dataset (209)")
print("  4. Zedda 2023 (YOLO-PAM) - ✅ IML (313) + MP-IDB (209)")
print("  5. Sukumarran 2024 (two-stage) - ✅ IML (313) + MP-IDB (209)")
print("  6. This Study 2025 - ✅ IML (313) + MP-IDB (418) + MD_2019 (883)")
print()
print("📌 Dataset column EMPHASIZED to show fair comparison!")
print("📌 All 5 papers use SAME datasets as ours (IML and/or MP-IDB)!")
