"""
Create Table: Dataset Statistics and Augmentation
Stacked-column format with wide columns and few rows
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
data_font = Font(size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "Dataset Augmentation"

# Headers (6 columns - stacked info in each)
headers = [
    'Dataset',
    'Original Split\n(Train/Val/Test)',
    'Original\nTotal',
    'Detection Augmentation\n(Train only, 4.4x)',
    'Classification Augmentation\n(Train only, 3.5x)',
    'Final Dataset Sizes\n(Detection / Classification)'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows - 4 datasets with stacked column format
data = [
    # Dataset, Original Split, Original Total, Detection Aug, Classification Aug, Final Sizes
    [
        'IML Lifecycle\n(4 stages)',
        '412 / 112 / 102',
        '626',
        'Train: 412→1,807\nVal: 112 (no aug)\nTest: 102 (no aug)\nTotal: 2,021',
        'Train: 412→1,446\nVal: 112 (no aug)\nTest: 102 (no aug)\nTotal: 1,660',
        'Detection: 2,021\nClassification: 1,660\nMultipliers: 4.4x / 3.5x'
    ],
    [
        'MP-IDB Species\n(4 species)',
        '274 / 72 / 72',
        '418',
        'Train: 274→1,202\nVal: 72 (no aug)\nTest: 72 (no aug)\nTotal: 1,346',
        'Train: 274→961\nVal: 72 (no aug)\nTest: 72 (no aug)\nTotal: 1,105',
        'Detection: 1,346\nClassification: 1,105\nMultipliers: 4.4x / 3.5x'
    ],
    [
        'MP-IDB Stages\n(4 stages)',
        '274 / 72 / 72',
        '418',
        'Train: 274→1,202\nVal: 72 (no aug)\nTest: 72 (no aug)\nTotal: 1,346',
        'Train: 274→961\nVal: 72 (no aug)\nTest: 72 (no aug)\nTotal: 1,105',
        'Detection: 1,346\nClassification: 1,105\nMultipliers: 4.4x / 3.5x'
    ],
    [
        'MD_2019 Stages\n(3 stages)',
        '1,028 / 270 / 328',
        '1,626',
        'Train: 1,028→4,510\nVal: 270 (no aug)\nTest: 328 (no aug)\nTotal: 5,108',
        'Train: 1,028→3,608\nVal: 270 (no aug)\nTest: 328 (no aug)\nTotal: 4,206',
        'Detection: 5,108\nClassification: 4,206\nMultipliers: 4.4x / 3.5x'
    ]
]

# Insert data rows
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border
        cell.font = data_font

        if col_idx == 1:  # Dataset name column
            cell.alignment = left_align
        else:
            cell.alignment = center_align

# Set column widths (wide columns for stacked content)
ws.column_dimensions['A'].width = 22  # Dataset
ws.column_dimensions['B'].width = 20  # Original Split
ws.column_dimensions['C'].width = 12  # Original Total
ws.column_dimensions['D'].width = 28  # Detection Aug
ws.column_dimensions['E'].width = 28  # Classification Aug
ws.column_dimensions['F'].width = 28  # Final Sizes

# Set row heights for readability
ws.row_dimensions[1].height = 45  # Header
for row in range(2, 6):  # Data rows
    ws.row_dimensions[row].height = 75  # Tall enough for multi-line content

# Save workbook
output_path = 'luaran/templates/tables/Table_Dataset_Augmentation.xlsx'
wb.save(output_path)

print("✅ Table: Dataset Statistics and Augmentation created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Table Structure:")
print("  - 4 datasets (IML, MP-IDB Species, MP-IDB Stages, MD_2019)")
print("  - 6 columns with stacked information")
print("  - Wide columns, few rows (stacked format)")
print()
print("📌 Key Information:")
print("  - Original 60/20/20 splits for all datasets")
print("  - Detection augmentation: 4.4x multiplier (train only)")
print("  - Classification augmentation: 3.5x multiplier (train only)")
print("  - Val/Test: NEVER augmented (fair evaluation)")
print()
print("💡 Placement in paper:")
print("  - Section 2.1 Datasets and Preprocessing")
print("  - After line 84 (augmentation description)")
print("  - Before Figure 1 (augmentation examples)")
