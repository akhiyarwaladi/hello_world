"""
Create Improved Table: Dataset Statistics and Augmentation
Multi-level headers with detailed Train/Val/Test breakdown
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define styles
header1_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header2_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
data_fill_alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header2_font = Font(bold=True, color="FFFFFF", size=10)
data_font = Font(size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "Dataset Augmentation"

# Level 1 Headers (merged cells)
# A1: Dataset
ws['A1'] = 'Dataset'
ws.merge_cells('A1:A2')
ws['A1'].fill = header1_fill
ws['A1'].font = header_font
ws['A1'].alignment = center_align
ws['A1'].border = border

# B1-E1: Original Data (60/20/20 Split)
ws['B1'] = 'Original Data (60/20/20 Split)'
ws.merge_cells('B1:E1')
ws['B1'].fill = header1_fill
ws['B1'].font = header_font
ws['B1'].alignment = center_align
ws['B1'].border = border

# F1-I1: Detection Training Data (4.4× Augmentation)
ws['F1'] = 'Detection Training Data\n(4.4× Augmentation on Train Only)'
ws.merge_cells('F1:I1')
ws['F1'].fill = header1_fill
ws['F1'].font = header_font
ws['F1'].alignment = center_align
ws['F1'].border = border

# J1-M1: Classification Training Data (3.5× Augmentation)
ws['J1'] = 'Classification Training Data\n(3.5× Augmentation on Train Only)'
ws.merge_cells('J1:M1')
ws['J1'].fill = header1_fill
ws['J1'].font = header_font
ws['J1'].alignment = center_align
ws['J1'].border = border

# Level 2 Headers
level2_headers = [
    # Original columns
    'Train', 'Val', 'Test', 'Total',
    # Detection columns
    'Train', 'Val', 'Test', 'Total',
    # Classification columns
    'Train', 'Val', 'Test', 'Total'
]

for col_idx, header in enumerate(level2_headers, 2):  # Start from column B (index 2)
    cell = ws.cell(2, col_idx)
    cell.value = header
    cell.fill = header2_fill
    cell.font = header2_font
    cell.alignment = center_align
    cell.border = border

# Data rows - from actual experiment results
data = [
    # [Dataset, Orig_Train, Orig_Val, Orig_Test, Orig_Total, Det_Train, Det_Val, Det_Test, Det_Total, Cls_Train, Cls_Val, Cls_Test, Cls_Total]
    [
        'IML Lifecycle\n(4 stages)',
        412, 112, 102, 626,
        1807, 112, 102, 2021,
        1446, 112, 102, 1660
    ],
    [
        'MP-IDB Species\n(4 species)',
        274, 72, 72, 418,
        1202, 72, 72, 1346,
        961, 72, 72, 1105
    ],
    [
        'MP-IDB Stages\n(4 stages)',
        274, 72, 72, 418,
        1202, 72, 72, 1346,
        961, 72, 72, 1105
    ],
    [
        'MD_2019 Stages\n(3 stages)',
        1028, 270, 328, 1626,
        4510, 270, 328, 5108,
        3608, 270, 328, 4206
    ]
]

# Insert data rows
for row_idx, row_data in enumerate(data, 3):  # Start from row 3
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border
        cell.font = data_font

        if col_idx == 1:  # Dataset name column
            cell.alignment = left_align
        else:
            cell.alignment = center_align

        # Alternating row colors for data
        if row_idx % 2 == 0:
            cell.fill = data_fill_alt

# Set column widths
ws.column_dimensions['A'].width = 20  # Dataset
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
    ws.column_dimensions[col].width = 10  # All data columns

# Set row heights
ws.row_dimensions[1].height = 40  # Header level 1
ws.row_dimensions[2].height = 25  # Header level 2
for row in range(3, 7):  # Data rows
    ws.row_dimensions[row].height = 35

# Save workbook
output_path = 'luaran/templates/tables/Table1_Dataset_Augmentation.xlsx'
wb.save(output_path)

print("✅ Table 1: Dataset Statistics and Augmentation created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Table Structure:")
print("  - Multi-level headers (2 levels)")
print("  - 13 columns: 1 dataset + 4 original + 4 detection + 4 classification")
print("  - 4 data rows (4 datasets)")
print()
print("📌 Header Levels:")
print("  Level 1: Dataset | Original Data | Detection Aug | Classification Aug")
print("  Level 2: Train, Val, Test, Total (for each group)")
print()
print("💡 Key Features:")
print("  - Clear separation of Train/Val/Test in separate columns")
print("  - Val/Test remain unchanged (112/102 for IML)")
print("  - Only Train augmented (412→1,807 detection, 412→1,446 classification)")
print("  - Totals calculated correctly")
print()
print("📊 Data Verified Against:")
print("  results/optA_20251016_200330/experiments/*/analysis_dataset_statistics/dataset_statistics_detailed.csv")
