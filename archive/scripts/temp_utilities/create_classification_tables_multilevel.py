"""
Create Classification Tables (2-6) with 2-level headers
Level 1: Model, Params, Accuracy, Bal.Acc, F1-Score (merged)
Level 2: Class names under F1-Score
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_classification_table(output_path, sheet_name, models_data, classes):
    """
    Create classification table with 2-level headers

    Args:
        output_path: Path to save Excel file
        sheet_name: Name of worksheet
        models_data: List of dicts with keys: model, params, accuracy, bal_acc, f1_scores
        classes: List of class names (with sample counts)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Calculate column positions
    num_classes = len(classes)
    f1_start_col = 5  # After Model, Params, Accuracy, Bal.Acc
    f1_end_col = f1_start_col + num_classes - 1

    # Row 1: Top-level headers
    # Model (A1:A2)
    ws.merge_cells('A1:A2')
    ws['A1'] = 'Model'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].border = border

    # Params (M) (B1:B2)
    ws.merge_cells('B1:B2')
    ws['B1'] = 'Params (M)'
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = center_align
    ws['B1'].border = border

    # Accuracy (%) (C1:C2)
    ws.merge_cells('C1:C2')
    ws['C1'] = 'Accuracy (%)'
    ws['C1'].fill = header_fill
    ws['C1'].font = header_font
    ws['C1'].alignment = center_align
    ws['C1'].border = border

    # Balanced Acc (%) (D1:D2)
    ws.merge_cells('D1:D2')
    ws['D1'] = 'Balanced Acc (%)'
    ws['D1'].fill = header_fill
    ws['D1'].font = header_font
    ws['D1'].alignment = center_align
    ws['D1'].border = border

    # F1-Score (merged across all classes)
    f1_start_letter = get_column_letter(f1_start_col)
    f1_end_letter = get_column_letter(f1_end_col)
    ws.merge_cells(f'{f1_start_letter}1:{f1_end_letter}1')
    ws[f'{f1_start_letter}1'] = 'F1-Score'
    ws[f'{f1_start_letter}1'].fill = header_fill
    ws[f'{f1_start_letter}1'].font = header_font
    ws[f'{f1_start_letter}1'].alignment = center_align
    ws[f'{f1_start_letter}1'].border = border

    # Row 2: Class names (subheaders under F1-Score)
    for i, class_name in enumerate(classes):
        col = f1_start_col + i
        cell = ws.cell(row=2, column=col)
        cell.value = class_name
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border

    # Apply border to all level 1 headers
    for col in range(1, f1_end_col + 1):
        ws.cell(row=1, column=col).border = border

    # Fill data rows
    for row_idx, model_data in enumerate(models_data):
        row = row_idx + 3

        # Model name
        cell = ws.cell(row=row, column=1)
        cell.value = model_data['model']
        cell.font = Font(bold=False, size=10)
        cell.alignment = left_align
        cell.border = border

        # Params
        cell = ws.cell(row=row, column=2)
        cell.value = model_data['params']
        cell.number_format = '0.0'
        cell.alignment = center_align
        cell.border = border

        # Accuracy
        cell = ws.cell(row=row, column=3)
        cell.value = model_data['accuracy']
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

        # Balanced Accuracy
        cell = ws.cell(row=row, column=4)
        cell.value = model_data['bal_acc']
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

        # F1-Scores
        for col_idx, f1_value in enumerate(model_data['f1_scores']):
            cell = ws.cell(row=row, column=f1_start_col + col_idx)
            cell.value = f1_value
            cell.number_format = '0.0000'
            cell.alignment = center_align
            cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    for col in range(f1_start_col, f1_end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Freeze top 2 rows
    ws.freeze_panes = 'A3'

    # Save
    wb.save(output_path)

# ============================================================================
# TABLE 2: IML Lifecycle Classification
# ============================================================================
def create_table2():
    models_data = [
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 91.51,
            'bal_acc': 91.96,
            'f1_scores': [0.9474, 0.9296, 0.8889, 0.8108]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 91.51,
            'bal_acc': 90.35,
            'f1_scores': [0.9592, 0.9429, 0.8889, 0.7429]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 91.51,
            'bal_acc': 90.57,
            'f1_scores': [0.9485, 0.9444, 1.0000, 0.7429]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 89.62,
            'bal_acc': 88.75,
            'f1_scores': [0.9474, 0.9067, 1.0000, 0.7059]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 87.74,
            'bal_acc': 87.86,
            'f1_scores': [0.9375, 0.9254, 0.8000, 0.6667]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 85.85,
            'bal_acc': 80.29,
            'f1_scores': [0.9278, 0.8857, 0.7500, 0.6486]
        }
    ]

    classes = ['Gametocyte (n=49)', 'Ring (n=34)', 'Schizont (n=4)', 'Trophozoite (n=19)']

    create_classification_table(
        'luaran/templates/tables/Table2_IML_Classification.xlsx',
        'IML Lifecycle',
        models_data,
        classes
    )
    print("✅ Table 2 created: IML Lifecycle Classification")

# ============================================================================
# TABLE 3: MP-IDB Species Classification
# ============================================================================
def create_table3():
    models_data = [
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 98.28,
            'bal_acc': 86.43,
            'f1_scores': [0.9942, 0.9333, 0.8000, 0.8571]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 97.93,
            'bal_acc': 80.95,
            'f1_scores': [0.9942, 0.9091, 0.8000, 0.7273]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 97.59,
            'bal_acc': 79.29,
            'f1_scores': [0.9942, 0.8750, 0.8000, 0.6667]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 97.59,
            'bal_acc': 83.54,
            'f1_scores': [0.9923, 0.9091, 0.7778, 0.7273]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 97.93,
            'bal_acc': 83.63,
            'f1_scores': [0.9942, 0.9091, 0.8235, 0.7273]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 97.24,
            'bal_acc': 79.19,
            'f1_scores': [0.9923, 0.8750, 0.7500, 0.6667]
        }
    ]

    classes = ['P.falciparum (n=259)', 'P.vivax (n=15)', 'P.malariae (n=9)', 'P.ovale (n=7)']

    create_classification_table(
        'luaran/templates/tables/Table3_Species_Classification.xlsx',
        'MP-IDB Species',
        models_data,
        classes
    )
    print("✅ Table 3 created: MP-IDB Species Classification")

# ============================================================================
# TABLE 4: MP-IDB Stages Classification
# ============================================================================
def create_table4():
    models_data = [
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 96.13,
            'bal_acc': 83.04,
            'f1_scores': [0.9846, 0.6087, 0.7143, 0.9091]
        },
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 95.42,
            'bal_acc': 78.64,
            'f1_scores': [0.9846, 0.4762, 0.7500, 0.7273]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 94.37,
            'bal_acc': 66.99,
            'f1_scores': [0.9827, 0.4800, 0.5000, 0.7500]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 94.72,
            'bal_acc': 70.31,
            'f1_scores': [0.9827, 0.4167, 0.6154, 0.8000]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 95.07,
            'bal_acc': 70.57,
            'f1_scores': [0.9827, 0.5714, 0.7273, 0.6000]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 92.25,
            'bal_acc': 77.91,
            'f1_scores': [0.9647, 0.4828, 0.4444, 0.9091]
        }
    ]

    classes = ['Ring (n=259)', 'Trophozoite (n=14)', 'Schizont (n=6)', 'Gametocyte (n=5)']

    create_classification_table(
        'luaran/templates/tables/Table4_Stages_Classification.xlsx',
        'MP-IDB Stages',
        models_data,
        classes
    )
    print("✅ Table 4 created: MP-IDB Stages Classification")

# ============================================================================
# TABLE 5: MD_2019 Classification
# ============================================================================
def create_table5():
    models_data = [
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 86.45,
            'bal_acc': 84.13,
            'f1_scores': [0.8864, 0.9184, 0.7120]
        },
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 85.25,
            'bal_acc': 83.20,
            'f1_scores': [0.8964, 0.8996, 0.6853]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 84.91,
            'bal_acc': 81.88,
            'f1_scores': [0.8850, 0.9031, 0.6747]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 84.56,
            'bal_acc': 83.70,
            'f1_scores': [0.8812, 0.8954, 0.7029]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 84.39,
            'bal_acc': 82.46,
            'f1_scores': [0.8852, 0.8905, 0.6825]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 84.22,
            'bal_acc': 81.36,
            'f1_scores': [0.8768, 0.9014, 0.6586]
        }
    ]

    classes = ['Ring (n=170)', 'Schizont (n=286)', 'Trophozoite (n=127)']

    create_classification_table(
        'luaran/templates/tables/Table5_MD2019_Classification.xlsx',
        'MD_2019 Stages',
        models_data,
        classes
    )
    print("✅ Table 5 created: MD_2019 Classification")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print("Creating Classification Tables with 2-level headers...")
    create_table2()
    create_table3()
    create_table4()
    create_table5()
    print("\n✅ All classification tables created successfully!")
    print("📁 Location: luaran/templates/tables/")
    print("\n📊 Structure:")
    print("   - Level 1: Model | Params | Accuracy | Bal.Acc | F1-Score (merged)")
    print("   - Level 2: Class names under F1-Score")
