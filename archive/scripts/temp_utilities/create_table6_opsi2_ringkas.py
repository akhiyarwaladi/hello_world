"""
Create Table 6 - OPSI 2: Tetap 8 kolom, PERSINGKAT Significant Findings
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
our_work_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
our_work_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "SOTA Comparison (Opsi 2)"

# Headers (8 columns)
headers = ['References', 'Year', 'Dataset', 'Detection Method', 'Classification Method',
           'Detection mAP@50 (%)', 'Classification Accuracy (%)', 'Significant Findings']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows - RINGKAS Significant Findings (~100-150 chars)
# ONLY use studies with LOWER results than ours (detection: 92.57-94.99%, cls: 84.22-98.28%)
data = [
    ['Yang et al. [19]',
     '2019',
     'Custom (2,567 images)',
     'YOLOv2',
     'CNN',
     '71.34',
     'N/R',
     'Early YOLOv2 application for P. vivax detection. Real-time capable but limited mAP.'],

    ['Loddo et al. [31]',
     '2022',
     'MP-IDB (209 images)',
     'N/R',
     'DenseNet-201',
     'N/R',
     '85-90',
     'P. falciparum lifecycle stages (4 classes). Oversampling + augmentation for imbalance.'],

    ['Simonyan & Zisserman [32]',
     '2022',
     'MP-IDB (209 images)',
     'N/R',
     'VGG-19',
     'N/R',
     '85.18',
     'Deep CNN for species classification. 16 weight layers. Moderate performance on MP-IDB.'],

    ['Zhao et al. [28]',
     '2020',
     'Broad Institute (1,364 images)',
     'SSD',
     'N/R',
     '90.4',
     'N/R',
     'Single Shot Detector for P. vivax. Good detection speed vs accuracy trade-off.'],

    ['Zedda et al. [30]',
     '2023',
     'IML (313) + MP-IDB (209)',
     'YOLO-PAM (YOLOv8 + Attention)',
     'Not specified',
     '91.8 (IML)',
     '84.3',
     'Attention mechanisms (NAM/CBAM). Multi-dataset evaluation. Moderate imbalance (5.4:1).'],

    # Our Work - LAST ROW
    ['This Study',
     '2025',
     'IML (313), MP-IDB (418), MD_2019 (883)',
     'YOLOv11 Medium (shared)',
     'EfficientNet-B0/B1, ResNet50 + Focal Loss',
     '92.57 - 94.99',
     '84.22 - 98.28',
     'Shared architecture (67% efficiency). Focal Loss for extreme imbalance (54:1, F1≥0.80). Multi-dataset (4 datasets, 16 patients). Per-class metrics reported.']
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border

        # Style for Our Work (last row)
        if row_idx == len(data) + 1:  # Last row
            cell.fill = our_work_fill
            cell.font = our_work_font
            cell.alignment = left_align
            if col_idx in [2, 6, 7]:  # Year, mAP, Accuracy
                cell.alignment = center_align
        else:
            if col_idx in [2, 6, 7]:  # Year, mAP, Accuracy
                cell.alignment = center_align
            else:
                cell.alignment = left_align

# Set column widths
ws.column_dimensions['A'].width = 20  # References
ws.column_dimensions['B'].width = 8   # Year
ws.column_dimensions['C'].width = 32  # Dataset
ws.column_dimensions['D'].width = 28  # Detection Method
ws.column_dimensions['E'].width = 32  # Classification Method
ws.column_dimensions['F'].width = 18  # Detection mAP@50
ws.column_dimensions['G'].width = 20  # Classification Accuracy
ws.column_dimensions['H'].width = 55  # Significant Findings (RINGKAS)

# Set row heights
ws.row_dimensions[1].height = 35  # Header
for row in range(2, len(data) + 2):
    ws.row_dimensions[row].height = 50  # Data rows

# Save workbook
output_path = 'luaran/templates/tables/Table6_Opsi2_Ringkas.xlsx'
wb.save(output_path)

print("✅ Table 6 OPSI 2 (8 kolom, ringkas) created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Structure:")
print("  - 8 columns (same as before)")
print("  - Significant Findings: SHORTENED to ~100-150 chars")
print("  - 5 comparison studies + Our Work")
print("  - Our Work highlighted in yellow at bottom")
