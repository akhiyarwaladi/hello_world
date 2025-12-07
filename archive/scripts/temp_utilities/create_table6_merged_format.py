"""
Create Table 6: SOTA Comparison (Merged Detection + Classification)
Format inspired by Tables 10 & 11 from reference paper
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
our_work_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
our_work_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "SOTA Comparison"

# Headers
headers = ['References', 'Year', 'Dataset', 'Detection Method', 'Classification Method',
           'Detection mAP@50 (%)', 'Classification Accuracy (%)', 'Significant Findings']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows - Other studies first
data = [
    ['Krishnadas et al. [29]',
     '2022',
     'Custom dataset (500 images)',
     'Faster R-CNN + ResNet50',
     'ResNet50',
     '89.2',
     '82.5',
     'Two-stage approach for parasite detection and species classification. Precision: 74.1%, Recall: 79.7% for species classification.'],

    ['Zedda et al. [30]',
     '2023',
     'IML Lifecycle (313 images)',
     'YOLO-PAM (YOLOv8 + Attention)',
     'Not specified',
     '84.6',
     '84.3',
     'Real-time detection with attention mechanisms. Lifecycle stage classification with moderate imbalance (5.4:1 ratio).'],

    ['Loddo et al. [31]',
     '2019',
     'MP-IDB (209 images)',
     'Mask R-CNN + DenseNet',
     'DenseNet',
     '88.7',
     '90.2',
     'Instance segmentation approach for precise localization. Focus on species classification (4 Plasmodium species).'],

    ['Chaudhry et al. [32]',
     '2024',
     'Mixed datasets (800 images)',
     'YOLOv8 + Vision Transformer',
     'Vision Transformer',
     '92.5',
     '88.6',
     'Attention mechanisms with multi-dataset training. Combined detection and classification in unified framework.'],

    ['Rajaraman et al. [33]',
     '2022',
     'NIH dataset (27,000 cells, balanced)',
     'N/A',
     'Ensemble CNNs',
     'N/A',
     '96.8',
     'Cell-level classification only (no detection). High accuracy achieved on artificially balanced dataset (1:1 ratio).'],

    # Our Work - LAST ROW
    ['This study (Our Work)',
     '2025',
     'Multi-dataset: IML Lifecycle (313), MP-IDB Species (209), MP-IDB Stages (209), MD_2019 (883) - Total: 1,614 images',
     'YOLOv11 Medium (shared across datasets)',
     'Shared architecture: EfficientNet-B0/B1, ResNet50 with Focal Loss (α=0.25, γ=2.0)',
     '92.57 - 94.99',
     '84.22 - 98.28',
     'First stage: YOLOv11 detection (mAP@50: 92.57-94.99%, Precision: 68.58-92.91%, Recall: 75.70-92.59%). Second stage: Dataset-specific classification with shared models (67% model reduction vs detector-specific). Focal Loss handles extreme imbalance (54:1 ratio) with F1-scores 0.80-1.00 on minority classes. Multi-dataset validation (4 datasets) with 16-patient diversity. Per-class precision reported for medical safety (false positive control).']
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border

        # Style for Our Work (last row)
        if row_idx == len(data) + 1:  # Last row
            cell.fill = our_work_fill
            cell.font = our_work_font
            cell.alignment = left_align
            # Center align for numeric columns
            if col_idx in [2, 6, 7]:  # Year, mAP, Accuracy
                cell.alignment = center_align
        else:
            # Center align for numeric columns
            if col_idx in [2, 6, 7]:  # Year, mAP, Accuracy
                cell.alignment = center_align
            else:
                cell.alignment = left_align

# Set column widths
ws.column_dimensions['A'].width = 25  # References
ws.column_dimensions['B'].width = 8   # Year
ws.column_dimensions['C'].width = 30  # Dataset
ws.column_dimensions['D'].width = 30  # Detection Method
ws.column_dimensions['E'].width = 35  # Classification Method
ws.column_dimensions['F'].width = 18  # Detection mAP@50
ws.column_dimensions['G'].width = 22  # Classification Accuracy
ws.column_dimensions['H'].width = 60  # Significant Findings

# Set row heights for better readability
for row in range(1, len(data) + 2):
    ws.row_dimensions[row].height = 60 if row > 1 else 30

# Save workbook
output_path = 'luaran/templates/tables/Table6_Comparison_SOTA.xlsx'
wb.save(output_path)

print("✅ Table 6: SOTA Comparison (Merged Format) created successfully!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Structure:")
print("  - Single sheet with merged detection + classification comparison")
print("  - 5 comparison studies + Our Work (1 row at bottom)")
print("  - Detailed 'Significant Findings' column")
print()
print("✨ Format matches reference Tables 10 & 11:")
print("  ✅ Simple reference format (e.g., 'Krishnadas et al. [29]')")
print("  ✅ Separate detection and classification method columns")
print("  ✅ Our Work at bottom (highlighted in yellow)")
print("  ✅ Detailed significant findings with metrics")
print("  ✅ Two-stage approach clearly shown")
