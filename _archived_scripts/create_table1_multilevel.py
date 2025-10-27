"""
Create Table 1 with 2-level headers (YOLO variants at top, metrics below)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Detection Results"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Row 1: Top-level headers (YOLO10, YOLO11, YOLO12)
ws.merge_cells('A1:A2')  # Dataset
ws['A1'] = 'Dataset'
ws['A1'].fill = header_fill
ws['A1'].font = header_font
ws['A1'].alignment = center_align
ws['A1'].border = border

# YOLO10 (B1:E1)
ws.merge_cells('B1:E1')
ws['B1'] = 'YOLOv10'
ws['B1'].fill = header_fill
ws['B1'].font = header_font
ws['B1'].alignment = center_align

# YOLO11 (F1:I1)
ws.merge_cells('F1:I1')
ws['F1'] = 'YOLOv11'
ws['F1'].fill = header_fill
ws['F1'].font = header_font
ws['F1'].alignment = center_align

# YOLO12 (J1:M1)
ws.merge_cells('J1:M1')
ws['J1'] = 'YOLOv12'
ws['J1'].fill = header_fill
ws['J1'].font = header_font
ws['J1'].alignment = center_align

# Row 2: Metric headers
metrics = ['mAP@50', 'mAP@50-95', 'Precision', 'Recall']
for i, metric in enumerate(metrics):
    # YOLO10 metrics
    cell = ws.cell(row=2, column=2+i)
    cell.value = metric
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = center_align
    cell.border = border

    # YOLO11 metrics
    cell = ws.cell(row=2, column=6+i)
    cell.value = metric
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = center_align
    cell.border = border

    # YOLO12 metrics
    cell = ws.cell(row=2, column=10+i)
    cell.value = metric
    cell.fill = subheader_fill
    cell.font = subheader_font
    cell.alignment = center_align
    cell.border = border

# Apply border to top headers
for col in range(1, 14):
    ws.cell(row=1, column=col).border = border
    ws.cell(row=2, column=col).border = border

# Data
datasets = ['IML Lifecycle', 'MP-IDB Species', 'MP-IDB Stages', 'MD_2019 Stages']

# YOLO10 data
yolo10_data = [
    [93.81, 77.71, 90.22, 92.22],  # IML
    [92.44, 60.12, 85.67, 90.73],  # Species
    [93.78, 44.48, 91.57, 93.12],  # Stages
    [70.84, 57.05, 67.89, 71.05]   # MD_2019
]

# YOLO11 data
yolo11_data = [
    [94.99, 77.76, 91.91, 91.11],  # IML
    [92.57, 62.17, 86.52, 91.88],  # Species
    [94.48, 60.34, 93.15, 88.36],  # Stages
    [72.91, 57.71, 68.58, 75.70]   # MD_2019
]

# YOLO12 data
yolo12_data = [
    [94.40, 78.21, 92.20, 86.67],  # IML
    [92.72, 62.25, 88.99, 89.28],  # Species
    [96.27, 61.53, 92.91, 92.59],  # Stages
    [71.12, 56.93, 65.92, 75.18]   # MD_2019
]

# Fill data rows
for row_idx, dataset in enumerate(datasets):
    row = row_idx + 3

    # Dataset name
    cell = ws.cell(row=row, column=1)
    cell.value = dataset
    cell.font = Font(bold=False, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

    # YOLO10 metrics
    for col_idx, value in enumerate(yolo10_data[row_idx]):
        cell = ws.cell(row=row, column=2+col_idx)
        cell.value = value
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

    # YOLO11 metrics
    for col_idx, value in enumerate(yolo11_data[row_idx]):
        cell = ws.cell(row=row, column=6+col_idx)
        cell.value = value
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

    # YOLO12 metrics
    for col_idx, value in enumerate(yolo12_data[row_idx]):
        cell = ws.cell(row=row, column=10+col_idx)
        cell.value = value
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

# Set column widths
ws.column_dimensions['A'].width = 18
for col in range(2, 14):
    ws.column_dimensions[get_column_letter(col)].width = 11

# Freeze top 2 rows
ws.freeze_panes = 'A3'

# Save
output_path = 'luaran/templates/tables/Table1_Detection_Performance.xlsx'
wb.save(output_path)

print("✅ Table 1 created with 2-level headers!")
print(f"📊 Structure:")
print(f"   - Level 1: Dataset | YOLOv10 | YOLOv11 | YOLOv12")
print(f"   - Level 2: Metrics (mAP@50, mAP@50-95, Precision, Recall)")
print(f"   - Data rows: 4 datasets")
print(f"   - Total: 6 rows × 13 columns")
print(f"\n✅ Saved to: {output_path}")
