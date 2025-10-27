"""
Create Classification Tables (2-5) with 2-level headers
- Format: Decimal 2 digits (0.XX)
- Structure: Per-class Precision + F1
- Level 1: Model, Params, Accuracy, Bal.Acc, Class1, Class2, ...
- Level 2: (under each class) Precision | F1
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
subheader_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_classification_table(output_path, sheet_name, models_data, classes_info):
    """
    Create classification table with 2-level headers grouped by class

    Args:
        output_path: Path to save Excel file
        sheet_name: Name of worksheet
        models_data: List of dicts with keys: model, params, accuracy, bal_acc, per_class_metrics
        classes_info: List of tuples (class_name, class_count)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    num_classes = len(classes_info)

    # Calculate column positions
    # Base: Model(1), Params(2), Accuracy(3), Bal.Acc(4)
    # Per-class: 2 metrics each (Precision, F1)
    base_cols = 4
    class_start_col = base_cols + 1

    # Row 1: Top-level headers
    # Model (A1:A2)
    ws.merge_cells('A1:A2')
    ws['A1'] = 'Model'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    ws['A1'].border = border

    # Params (B1:B2)
    ws.merge_cells('B1:B2')
    ws['B1'] = 'Params (M)'
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = center_align
    ws['B1'].border = border

    # Accuracy (C1:C2)
    ws.merge_cells('C1:C2')
    ws['C1'] = 'Accuracy'
    ws['C1'].fill = header_fill
    ws['C1'].font = header_font
    ws['C1'].alignment = center_align
    ws['C1'].border = border

    # Balanced Acc (D1:D2)
    ws.merge_cells('D1:D2')
    ws['D1'] = 'Balanced Acc'
    ws['D1'].fill = header_fill
    ws['D1'].font = header_font
    ws['D1'].alignment = center_align
    ws['D1'].border = border

    # Class headers (each class gets 2 columns: Precision, F1)
    for i, (class_name, class_count) in enumerate(classes_info):
        col_start = class_start_col + (i * 2)
        col_end = col_start + 1

        # Merge 2 columns for class name
        start_letter = get_column_letter(col_start)
        end_letter = get_column_letter(col_end)
        ws.merge_cells(f'{start_letter}1:{end_letter}1')

        # Set class name with count
        cell = ws[f'{start_letter}1']
        cell.value = f'{class_name} (n={class_count})'
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Row 2: Metric subheaders (Precision, F1 under each class)
    for i in range(num_classes):
        col_start = class_start_col + (i * 2)

        # Precision
        cell = ws.cell(row=2, column=col_start)
        cell.value = 'Precision'
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border

        # F1
        cell = ws.cell(row=2, column=col_start + 1)
        cell.value = 'F1'
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = border

    # Apply border to all headers
    total_cols = base_cols + (num_classes * 2)
    for col in range(1, total_cols + 1):
        ws.cell(row=1, column=col).border = border
        ws.cell(row=2, column=col).border = border

    # Fill data rows
    for row_idx, model_data in enumerate(models_data):
        row = row_idx + 3

        # Model name
        cell = ws.cell(row=row, column=1)
        cell.value = model_data['model']
        cell.font = Font(bold=False, size=10)
        cell.alignment = left_align
        cell.border = border

        # Params
        cell = ws.cell(row=row, column=2)
        cell.value = model_data['params']
        cell.number_format = '0.0'
        cell.alignment = center_align
        cell.border = border

        # Accuracy (as decimal 0.XX)
        cell = ws.cell(row=row, column=3)
        cell.value = model_data['accuracy'] / 100.0  # Convert from percentage to decimal
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

        # Balanced Accuracy (as decimal 0.XX)
        cell = ws.cell(row=row, column=4)
        cell.value = model_data['bal_acc'] / 100.0  # Convert from percentage to decimal
        cell.number_format = '0.00'
        cell.alignment = center_align
        cell.border = border

        # Per-class metrics (Precision, F1)
        for class_idx, (prec, f1) in enumerate(model_data['per_class_metrics']):
            col_start = class_start_col + (class_idx * 2)

            # Precision
            cell = ws.cell(row=row, column=col_start)
            cell.value = prec
            cell.number_format = '0.00'
            cell.alignment = center_align
            cell.border = border

            # F1
            cell = ws.cell(row=row, column=col_start + 1)
            cell.value = f1
            cell.number_format = '0.00'
            cell.alignment = center_align
            cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 13
    for col in range(class_start_col, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Freeze top 2 rows
    ws.freeze_panes = 'A3'

    # Save
    wb.save(output_path)

# ============================================================================
# TABLE 2: IML Lifecycle Classification
# ============================================================================
def create_table2():
    # Data from table9_focal_loss.csv
    # Order: densenet121, efficientnet_b0, efficientnet_b1, efficientnet_b2, resnet101, resnet50
    # But we want to show best models first

    models_data = [
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 91.51,
            'bal_acc': 91.96,
            'per_class_metrics': [
                (0.9783, 0.9474),  # gametocyte: precision, f1
                (0.8919, 0.9296),  # ring
                (0.8000, 0.8889),  # schizont
                (0.8333, 0.8108),  # trophozoite
            ]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 91.51,
            'bal_acc': 90.35,
            'per_class_metrics': [
                (0.9592, 0.9592),  # gametocyte
                (0.9167, 0.9429),  # ring
                (0.8000, 0.8889),  # schizont
                (0.8125, 0.7429),  # trophozoite
            ]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 91.51,
            'bal_acc': 90.57,
            'per_class_metrics': [
                (0.9583, 0.9485),  # gametocyte
                (0.8947, 0.9444),  # ring
                (1.0000, 1.0000),  # schizont
                (0.8125, 0.7429),  # trophozoite
            ]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 89.62,
            'bal_acc': 88.75,
            'per_class_metrics': [
                (0.9783, 0.9474),  # gametocyte
                (0.8293, 0.9067),  # ring
                (1.0000, 1.0000),  # schizont
                (0.8000, 0.7059),  # trophozoite
            ]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 87.74,
            'bal_acc': 87.86,
            'per_class_metrics': [
                (0.9574, 0.9375),  # gametocyte
                (0.9394, 0.9254),  # ring
                (0.6667, 0.8000),  # schizont
                (0.6500, 0.6667),  # trophozoite
            ]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 85.85,
            'bal_acc': 80.29,
            'per_class_metrics': [
                (0.9375, 0.9278),  # gametocyte
                (0.8611, 0.8857),  # ring
                (0.7500, 0.7500),  # schizont
                (0.6667, 0.6486),  # trophozoite
            ]
        }
    ]

    classes_info = [
        ('Gametocyte', 49),
        ('Ring', 34),
        ('Schizont', 4),
        ('Trophozoite', 19)
    ]

    create_classification_table(
        'luaran/templates/tables/Table2_IML_Classification.xlsx',
        'IML Lifecycle',
        models_data,
        classes_info
    )
    print("✅ Table 2 created: IML Lifecycle Classification")

# ============================================================================
# TABLE 3: MP-IDB Species Classification
# ============================================================================
def create_table3():
    models_data = [
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 98.28,
            'bal_acc': 86.43,
            'per_class_metrics': [
                (0.9885, 0.9942),  # P_falciparum
                (0.9333, 0.9333),  # P_vivax
                (1.0000, 0.8000),  # P_malariae
                (0.8571, 0.8571),  # P_ovale
            ]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 97.93,
            'bal_acc': 80.95,
            'per_class_metrics': [
                (0.9885, 0.9942),  # P_falciparum
                (0.8333, 0.9091),  # P_vivax
                (1.0000, 0.8000),  # P_malariae
                (1.0000, 0.7273),  # P_ovale
            ]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 97.59,
            'bal_acc': 79.29,
            'per_class_metrics': [
                (0.9885, 0.9942),  # P_falciparum
                (0.8235, 0.8750),  # P_vivax
                (1.0000, 0.8000),  # P_malariae
                (0.8000, 0.6667),  # P_ovale
            ]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 97.59,
            'bal_acc': 83.54,
            'per_class_metrics': [
                (0.9923, 0.9923),  # P_falciparum
                (0.8333, 0.9091),  # P_vivax
                (0.7778, 0.7778),  # P_malariae
                (1.0000, 0.7273),  # P_ovale
            ]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 97.93,
            'bal_acc': 83.63,
            'per_class_metrics': [
                (0.9923, 0.9942),  # P_falciparum
                (0.8333, 0.9091),  # P_vivax
                (0.8750, 0.8235),  # P_malariae
                (1.0000, 0.7273),  # P_ovale
            ]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 97.24,
            'bal_acc': 79.19,
            'per_class_metrics': [
                (0.9885, 0.9923),  # P_falciparum
                (0.8235, 0.8750),  # P_vivax
                (0.8571, 0.7500),  # P_malariae
                (0.8000, 0.6667),  # P_ovale
            ]
        }
    ]

    classes_info = [
        ('P.falciparum', 259),
        ('P.vivax', 15),
        ('P.malariae', 9),
        ('P.ovale', 7)
    ]

    create_classification_table(
        'luaran/templates/tables/Table3_Species_Classification.xlsx',
        'MP-IDB Species',
        models_data,
        classes_info
    )
    print("✅ Table 3 created: MP-IDB Species Classification")

# ============================================================================
# TABLE 4: MP-IDB Stages Classification
# ============================================================================
def create_table4():
    models_data = [
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 96.13,
            'bal_acc': 83.04,
            'per_class_metrics': [
                (0.9808, 0.9846),  # ring
                (0.7778, 0.6087),  # trophozoite
                (0.6250, 0.7143),  # schizont
                (0.8333, 0.9091),  # gametocyte
            ]
        },
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 95.42,
            'bal_acc': 78.64,
            'per_class_metrics': [
                (0.9808, 0.9846),  # ring
                (0.7143, 0.4762),  # trophozoite
                (0.6000, 0.7500),  # schizont
                (0.6667, 0.7273),  # gametocyte
            ]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 94.37,
            'bal_acc': 66.99,
            'per_class_metrics': [
                (0.9808, 0.9827),  # ring
                (0.5455, 0.4800),  # trophozoite
                (0.4000, 0.5000),  # schizont
                (1.0000, 0.7500),  # gametocyte
            ]
        },
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 94.72,
            'bal_acc': 70.31,
            'per_class_metrics': [
                (0.9771, 0.9827),  # ring
                (0.5000, 0.4167),  # trophozoite
                (0.5714, 0.6154),  # schizont
                (0.8000, 0.8000),  # gametocyte
            ]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 95.07,
            'bal_acc': 70.57,
            'per_class_metrics': [
                (0.9808, 0.9827),  # ring
                (0.5714, 0.5714),  # trophozoite
                (0.8000, 0.7273),  # schizont
                (0.6000, 0.6000),  # gametocyte
            ]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 92.25,
            'bal_acc': 77.91,
            'per_class_metrics': [
                (0.9801, 0.9647),  # ring
                (0.4667, 0.4828),  # trophozoite
                (0.3333, 0.4444),  # schizont
                (0.8333, 0.9091),  # gametocyte
            ]
        }
    ]

    classes_info = [
        ('Ring', 259),
        ('Trophozoite', 14),
        ('Schizont', 6),
        ('Gametocyte', 5)
    ]

    create_classification_table(
        'luaran/templates/tables/Table4_Stages_Classification.xlsx',
        'MP-IDB Stages',
        models_data,
        classes_info
    )
    print("✅ Table 4 created: MP-IDB Stages Classification")

# ============================================================================
# TABLE 5: MD_2019 Classification
# ============================================================================
def create_table5():
    models_data = [
        {
            'model': 'EfficientNet-B0',
            'params': 5.3,
            'accuracy': 86.45,
            'bal_acc': 84.13,
            'per_class_metrics': [
                (0.8571, 0.8864),  # ring
                (0.9317, 0.9184),  # schizont
                (0.7236, 0.7120),  # trophozoite
            ]
        },
        {
            'model': 'EfficientNet-B1',
            'params': 7.8,
            'accuracy': 85.25,
            'bal_acc': 83.20,
            'per_class_metrics': [
                (0.8556, 0.8964),  # ring
                (0.9228, 0.8996),  # schizont
                (0.6935, 0.6853),  # trophozoite
            ]
        },
        {
            'model': 'EfficientNet-B2',
            'params': 9.2,
            'accuracy': 84.91,
            'bal_acc': 81.88,
            'per_class_metrics': [
                (0.8876, 0.8850),  # ring
                (0.8938, 0.9031),  # schizont
                (0.6885, 0.6747),  # trophozoite
            ]
        },
        {
            'model': 'DenseNet121',
            'params': 8.0,
            'accuracy': 84.56,
            'bal_acc': 83.70,
            'per_class_metrics': [
                (0.8686, 0.8812),  # ring
                (0.9421, 0.8954),  # schizont
                (0.6510, 0.7029),  # trophozoite
            ]
        },
        {
            'model': 'ResNet50',
            'params': 25.6,
            'accuracy': 84.39,
            'bal_acc': 82.46,
            'per_class_metrics': [
                (0.8449, 0.8852),  # ring
                (0.9151, 0.8905),  # schizont
                (0.6880, 0.6825),  # trophozoite
            ]
        },
        {
            'model': 'ResNet101',
            'params': 44.5,
            'accuracy': 84.22,
            'bal_acc': 81.36,
            'per_class_metrics': [
                (0.8547, 0.8768),  # ring
                (0.9078, 0.9014),  # schizont
                (0.6721, 0.6586),  # trophozoite
            ]
        }
    ]

    classes_info = [
        ('Ring', 170),
        ('Schizont', 286),
        ('Trophozoite', 127)
    ]

    create_classification_table(
        'luaran/templates/tables/Table5_MD2019_Classification.xlsx',
        'MD_2019 Stages',
        models_data,
        classes_info
    )
    print("✅ Table 5 created: MD_2019 Classification")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print("Creating Classification Tables with Precision + F1...")
    print("Format: Decimal 0.XX for all metrics\n")
    create_table2()
    create_table3()
    create_table4()
    create_table5()
    print("\n✅ All classification tables created successfully!")
    print("📁 Location: luaran/templates/tables/")
    print("\n📊 Structure:")
    print("   - Level 1: Model | Params | Accuracy | Bal.Acc | Class1 | Class2 | ...")
    print("   - Level 2: (under each class) Precision | F1")
    print("   - Format: All metrics in decimal 0.XX")
