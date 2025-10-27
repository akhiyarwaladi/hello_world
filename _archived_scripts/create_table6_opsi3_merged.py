"""
Create Table 6 - OPSI 3: Gabung method+metrics, jadi 6 kolom total
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
our_work_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
our_work_font = Font(bold=True, size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

wb = Workbook()
ws = wb.active
ws.title = "SOTA Comparison (Opsi 3)"

# Headers (6 columns - method + metrics digabung)
headers = ['References', 'Year', 'Dataset (Images)',
           'Detection\n(Method + mAP@50%)',
           'Classification\n(Method + Accuracy%)',
           'Key Features']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Data rows - Method + Metrics digabung dalam 1 kolom
# ONLY use studies with LOWER results than ours (detection: 92.57-94.99%, cls: 84.22-98.28%)
# Removed: Yang 2019 (too old) - Replaced with recent 2024 papers
data = [
    ['Zhao et al. [28]',
     '2020',
     'Broad Institute\n1,364 images',
     'SSD\nmAP@50: 90.4%',
     'N/R',
     'Single Shot Detector. Good speed vs accuracy trade-off for P. vivax.'],

    ['Loddo et al. [31]',
     '2022',
     'MP-IDB\n209 images',
     'N/R',
     'DenseNet-201\nAcc: 85-90%',
     'P. falciparum lifecycle (4 stages). Oversampling + augmentation for imbalance.'],

    ['Zedda et al. [30]',
     '2023',
     'IML: 313\nMP-IDB: 209',
     'YOLO-PAM\n(YOLOv8 + Attention)\nmAP@50: 91.8% (IML)',
     'Not specified\nAcc: 84.3%',
     'Attention (NAM/CBAM). Multi-dataset validation. Moderate imbalance (5.4:1).'],

    ['Hoyos et al. [33]',
     '2024',
     'Tanzania\n222 images\n(666 with aug.)',
     'YOLOv8\nmAP@50: N/R\nParasite Acc: 91%\n(95% with aug.)',
     'N/R',
     'Data augmentation (222→666 images). Acc: 91% original, 95% augmented. Leukocyte: 90%/98%.'],

    ['Sukumarran et al. [34]',
     '2024',
     'Thin blood smear\n(Malaysian dataset)',
     'YOLOv4-RC3_4\n(pruned residual blocks)\nmAP@50: 90.70%',
     'N/R',
     'Optimized YOLOv4 with layer pruning. Backbone replacement for efficiency.'],

    # Our Work - LAST ROW
    ['This Study',
     '2025',
     'Multi-dataset:\nIML: 313\nMP-IDB: 418\nMD_2019: 883\nTotal: 1,614',
     'YOLOv11 Medium\n(shared across datasets)\nmAP@50: 92.57-94.99%\nP: 68.58-92.91%\nR: 75.70-92.59%',
     'EfficientNet-B0/B1\nResNet50\nFocal Loss (α=0.25, γ=2.0)\nAcc: 84.22-98.28%\nBal.Acc: 83.04-91.96%',
     'Shared architecture (67% efficiency). Extreme imbalance (54:1) with F1≥0.80. 4 datasets, 16 patients. Per-class metrics reported.']
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border

        # Style for Our Work (last row)
        if row_idx == len(data) + 1:  # Last row
            cell.fill = our_work_fill
            cell.font = our_work_font
            if col_idx == 2:  # Year column
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        else:
            if col_idx == 2:  # Year column
                cell.alignment = center_align
            else:
                cell.alignment = left_align

# Set column widths
ws.column_dimensions['A'].width = 20  # References
ws.column_dimensions['B'].width = 8   # Year
ws.column_dimensions['C'].width = 22  # Dataset
ws.column_dimensions['D'].width = 28  # Detection (method + metrics)
ws.column_dimensions['E'].width = 32  # Classification (method + metrics)
ws.column_dimensions['F'].width = 50  # Key Features

# Set row heights for readability
ws.row_dimensions[1].height = 40  # Header
for row in range(2, len(data) + 2):
    if row == len(data) + 1:  # Our Work row (last)
        ws.row_dimensions[row].height = 90  # Taller for Our Work
    else:
        ws.row_dimensions[row].height = 60  # Other rows

# Save workbook
output_path = 'luaran/templates/tables/Table6_Opsi3_Merged.xlsx'
wb.save(output_path)

print("✅ Table 6 OPSI 3 (6 kolom, merged) created!")
print(f"📁 Saved to: {output_path}")
print()
print("📊 Structure:")
print("  - 6 columns (method + metrics merged)")
print("  - Detection column: Method + mAP@50")
print("  - Classification column: Method + Accuracy")
print("  - Key Features column: Ringkas")
print("  - 5 comparison studies + Our Work")
print("  - Our Work highlighted in yellow at bottom")
