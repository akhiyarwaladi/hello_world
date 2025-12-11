"""
Generate Classification Performance Tables from Experiment Results
==================================================================

Reads data dynamically from a specified experiment folder and generates
classification tables in KINETIK journal format.

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import sys
import re
import argparse

# Professional color scheme
COLORS = {
    'header_bg': '3C7A8C',
    'header_text': 'FFFFFF',
    'alt_row': 'F0F4F8',
    'border': 'CBD5E0',
}

def extract_training_time(results_file):
    """Extract training time from results.txt file"""
    try:
        with open(results_file, 'r') as f:
            content = f.read()
            match = re.search(r'Training Time:\s*([\d.]+)\s*min', content)
            if match:
                return float(match.group(1))
    except Exception as e:
        print(f"  ⚠ Could not read training time from {results_file}: {e}")
    return None

def read_dataset_data(experiment_folder, dataset_key):
    """Read classification data for a specific dataset"""
    exp_path = experiment_folder / 'experiments' / f'experiment_{dataset_key}'

    # Read table9_focal_loss.csv
    table9_path = exp_path / 'table9_focal_loss.csv'
    if not table9_path.exists():
        print(f"  ⚠ table9 not found: {table9_path}")
        return None

    df = pd.read_csv(table9_path)

    # Extract overall metrics
    overall = df[df['Class'] == 'Overall']
    if overall.empty:
        print(f"  ⚠ No overall metrics found")
        return None

    # Get model columns
    models = [col for col in df.columns if col not in ['Class', 'Metric']]

    # Get classes (exclude Overall)
    classes = df[df['Class'] != 'Overall']['Class'].unique().tolist()

    # Build data structure
    models_data = []
    for model in models:
        # Get overall metrics
        accuracy = overall[overall['Metric'] == 'accuracy'][model].values[0]
        balanced_acc = overall[overall['Metric'] == 'balanced_accuracy'][model].values[0]

        # Get training time
        results_path = exp_path / f'cls_{model}_focal' / 'results.txt'
        training_time = extract_training_time(results_path)

        # Get per-class metrics
        per_class_metrics = []
        for cls in classes:
            cls_data = df[df['Class'] == cls]
            precision = cls_data[cls_data['Metric'] == 'precision'][model].values[0]
            f1 = cls_data[cls_data['Metric'] == 'f1_score'][model].values[0]
            support = int(cls_data[cls_data['Metric'] == 'support'][model].values[0])
            per_class_metrics.append({
                'class': cls,
                'precision': precision * 100,  # Convert to percentage
                'f1': f1 * 100,  # Convert to percentage
                'support': support
            })

        # Get model parameters
        param_map = {
            'densenet121': 8.0,
            'efficientnet_b0': 5.3,
            'efficientnet_b1': 7.8,
            'efficientnet_b2': 9.1,
            'resnet50': 25.6,
            'resnet101': 44.5
        }

        models_data.append({
            'model': model,
            'params': param_map.get(model, 0),
            'training_time': training_time,
            'accuracy': accuracy * 100,  # Convert to percentage
            'balanced_acc': balanced_acc * 100,  # Convert to percentage
            'per_class_metrics': per_class_metrics
        })

    return {
        'models_data': models_data,
        'classes': classes
    }

def create_classification_table(output_path, sheet_name, dataset_data):
    """Create classification table with 2-level headers"""
    models_data = dataset_data['models_data']
    classes = dataset_data['classes']

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header row 1: Base headers + class names (merged across 2 columns)
    row1 = ['Model', 'Parameters (M)', 'Training Time (min)', 'Accuracy', 'Balanced Acc']
    ws.append(row1 + [''] * (len(classes) * 2))

    # Merge cells for class names
    start_col = 6
    for i, cls_info in enumerate([m for m in models_data[0]['per_class_metrics']]):
        cls = cls_info['class']
        support = cls_info['support']
        col = start_col + i * 2
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col)
        cell.value = f"{cls.replace('_', ' ').title()} (n={support})"
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Header row 2: Metric names
    row2 = [''] * 5  # Empty for base columns
    for _ in classes:
        row2.extend(['Precision', 'F1'])
    ws.append(row2)

    # Data rows
    for model_data in models_data:
        row = [
            model_data['model'].replace('_', '-').upper(),
            model_data['params'],
            model_data['training_time'] if model_data['training_time'] else '',
            model_data['accuracy'],
            model_data['balanced_acc']
        ]

        # Add per-class metrics
        for cls_metric in model_data['per_class_metrics']:
            row.extend([cls_metric['precision'], cls_metric['f1']])

        ws.append(row)

    # Apply styling
    apply_professional_styling(ws, num_classes=len(classes))

    # Save
    wb.save(output_path)
    print(f"  ✓ Created: {output_path.name}")

def apply_professional_styling(ws, num_classes):
    """Apply professional styling to worksheet"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    # Style header rows (rows 1-2)
    for row in [1, 2]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Style data rows
    for row in range(3, ws.max_row + 1):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border

            # Alignment
            if col == 1:  # Model name
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Number formatting (2 decimal places for all)
            if isinstance(cell.value, (int, float)) and col > 1:
                if col in [2, 3]:  # Parameters and Training Time
                    cell.number_format = '0.0'
                else:  # All metrics (2 decimals only)
                    cell.number_format = '0.00'

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20  # Model
    ws.column_dimensions['B'].width = 16  # Parameters
    ws.column_dimensions['C'].width = 18  # Training Time
    ws.column_dimensions['D'].width = 12  # Accuracy
    ws.column_dimensions['E'].width = 14  # Balanced Acc

    # Per-class columns
    for i in range(num_classes * 2):
        col_letter = get_column_letter(6 + i)
        ws.column_dimensions[col_letter].width = 12

    # Freeze headers
    ws.freeze_panes = 'A3'

def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description='Generate classification tables from experiment results')
    parser.add_argument('--experiment', type=str, default='optA_20251207_233941',
                       help='Experiment folder name (default: optA_20251207_233941)')
    parser.add_argument('--output', type=str, default=None,
                       help='Output folder (default: luaran/laporan_akhir/tables)')

    args = parser.parse_args()

    print("=" * 70)
    print("CLASSIFICATION TABLE GENERATOR FROM EXPERIMENT")
    print("=" * 70)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent
        experiment_folder = base_dir / 'results' / args.experiment

        if not experiment_folder.exists():
            print(f"\n❌ Experiment folder not found: {experiment_folder}")
            sys.exit(1)

        if args.output:
            output_folder = Path(args.output)
        else:
            output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'

        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}\n")

        datasets = {
            'iml_lifecycle': ('IML Lifecycle', 'Table3'),
            'mp_idb_species': ('MP-IDB Species', 'Table4'),
            'mp_idb_stages': ('MP-IDB Stages', 'Table5'),
            'md_2019_stages': ('MD-2019 Stages', 'Table6')
        }

        for dataset_key, (dataset_name, table_name) in datasets.items():
            print(f"Processing {dataset_name}...")

            dataset_data = read_dataset_data(experiment_folder, dataset_key)
            if dataset_data is None:
                continue

            output_path = output_folder / f'{table_name}_{dataset_key}.xlsx'
            create_classification_table(output_path, dataset_name, dataset_data)

        print("\n" + "=" * 70)
        print("ALL CLASSIFICATION TABLES GENERATED SUCCESSFULLY!")
        print("=" * 70)
        print(f"\n📁 Tables saved to: {output_folder}")
        print(f"📊 Source experiment: {args.experiment}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
