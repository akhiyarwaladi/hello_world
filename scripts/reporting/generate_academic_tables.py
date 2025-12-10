"""
Generate Academic-Style Excel Tables for Final Report
=====================================================

Follows journal table format from KINETIK Paper:
- NO background colors (clean white)
- Simple black borders
- Bold headers only
- Minimal professional academic style
- Only 3 consolidated files

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

def apply_academic_styling(ws):
    """
    Apply simple academic journal table styling
    - No background colors
    - Simple borders
    - Bold headers
    """
    # Simple thin black border
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Header: bold, centered, no color
    header_font = Font(bold=True, size=11, name='Arial')

    # Apply header styling (first row)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows: simple borders, centered numbers
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # Left align for first column (names), center for numbers
            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Format numbers
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'

    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

def generate_dataset_statistics(experiment_folder, output_folder):
    """Generate dataset statistics table - KINETIK Table 1 format"""
    print("Generating Table 1: Dataset Statistics (KINETIK Format)...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'dataset_statistics_all.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    # Remove duplicates and get unique rows
    df = df.drop_duplicates(subset=['Dataset'])

    # Rename dataset for display
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle\n(4 stages)',
        'mp_idb_species': 'MP-IDB Species\n(4 species)',
        'mp_idb_stages': 'MP-IDB Stages\n(4 stages)',
        'md_2019_stages': 'MD-2019 Stages\n(3 stages)'
    }
    df['Dataset'] = df['Dataset'].map(dataset_names)

    # Restructure columns to match KINETIK Table 1 format
    # Detection (Full Images) | Classification (Bounding Box Crops) | After Augmentation | Boxes per Image
    result_df = pd.DataFrame({
        'Dataset': df['Dataset'],
        'Detection\nTrain': df['Original_Train'],
        'Detection\nVal': df['Original_Val'],
        'Detection\nTest': df['Original_Test'],
        'Detection\nTotal': df['Original_Train'] + df['Original_Val'] + df['Original_Test'],
        'Classification\nTrain': df['Original_Train'],  # Same as detection
        'Classification\nVal': df['Original_Val'],
        'Classification\nTest': df['Original_Test'],
        'Classification\nTotal': df['Original_Train'] + df['Original_Val'] + df['Original_Test'],
        'Aug Det Train': df['Detection_Multiplier'],
        'Aug Cls Train': df['Classification_Multiplier'],
        'Aug Val\n(same)': 'same',
        'Aug Test\n(same)': 'same',
        'Boxes per\nImage': '2.0×'  # From paper
    })

    output_file = output_folder / 'Table1_Dataset_Statistics.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='Dataset Statistics', index=False)
        apply_academic_styling(writer.sheets['Dataset Statistics'])

    print(f"  ✓ Created: {output_file.name}")

def generate_detection_performance(experiment_folder, output_folder):
    """Generate detection performance table (KINETIK format: models as columns per dataset)"""
    print("\nGenerating Table 2: Detection Performance (KINETIK Format)...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    # Map dataset names
    dataset_names = {
        'iml_lifecycle': 'IML Malaria\n(4 stages)',
        'mp_idb_species': 'MP-IDB\n(4 species)',
        'mp_idb_stages': 'MP-IDB\n(4 stages)',
        'md_2019_stages': 'MD-2019\n(3 stages)'
    }

    # Pivot to KINETIK format: one row per dataset, columns grouped by YOLO model
    result_rows = []

    for dataset_key, dataset_name in dataset_names.items():
        dataset_df = df[df['Dataset'] == dataset_key]

        if dataset_df.empty:
            continue

        row = {'Dataset': dataset_name}

        # For each YOLO model (UPPERCASE in CSV: YOLO10, YOLO11, YOLO12)
        for model_name in ['YOLO10', 'YOLO11', 'YOLO12']:
            model_df = dataset_df[dataset_df['Model'] == model_name]

            if not model_df.empty:
                model_data = model_df.iloc[0]

                # Get training time from detection results.csv
                training_time = 0
                det_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}' / f'det_{model_name.lower()}'
                results_csv = det_folder / 'results.csv'
                if results_csv.exists():
                    try:
                        det_df = pd.read_csv(results_csv)
                        if 'time' in det_df.columns and len(det_df) > 0:
                            # Last row has total training time in seconds
                            total_seconds = det_df['time'].iloc[-1]
                            training_time = round(total_seconds / 60, 2)  # Convert to minutes
                    except:
                        training_time = 0

                # Add columns for this model - use YOLOv10/v11/v12 format like in KINETIK
                yolo_label = f"YOLOv{model_name[-2:]}"  # YOLO10 -> YOLOv10
                row[f'{yolo_label}\nmAP@50'] = round(model_data['mAP@50'] * 100, 2)
                row[f'{yolo_label}\nmAP@50-95'] = round(model_data['mAP@50-95'] * 100, 2)
                row[f'{yolo_label}\nPrecision'] = round(model_data['Precision'] * 100, 2)
                row[f'{yolo_label}\nRecall'] = round(model_data['Recall'] * 100, 2)
                row[f'{yolo_label}\nTime(min)'] = training_time

        result_rows.append(row)

    result_df = pd.DataFrame(result_rows)

    output_file = output_folder / 'Table2_Detection_Performance.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='Detection Performance', index=False)
        apply_academic_styling(writer.sheets['Detection Performance'])

    print(f"  ✓ Created: {output_file.name}")
    print(f"    - {len(result_df)} rows (4 datasets × 3 YOLO models)")

def generate_classification_performance(experiment_folder, output_folder):
    """Generate classification performance tables (KINETIK format: one table per dataset)"""
    print("\nGenerating Tables 3-6: Classification Performance (KINETIK Format)...")

    # Model parameters (M) from architecture specs
    model_params = {
        'densenet121': 8.0,
        'efficientnet_b0': 5.3,
        'efficientnet_b1': 7.8,
        'efficientnet_b2': 9.2,
        'resnet50': 25.6,
        'resnet101': 44.5
    }

    datasets = {
        'iml_lifecycle': ('IML Lifecycle', 'Table3'),
        'mp_idb_species': ('MP-IDB Species', 'Table4'),
        'mp_idb_stages': ('MP-IDB Stages', 'Table5'),
        'md_2019_stages': ('MD-2019 Stages', 'Table6')
    }

    table_count = 0

    for dataset_key, (dataset_name, table_num) in datasets.items():
        exp_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}'
        table_file = exp_folder / 'table9_focal_loss.csv'

        if not table_file.exists():
            print(f"  ⚠ Not found for {dataset_name}: {table_file}")
            continue

        # Read pivot table
        df = pd.read_csv(table_file)
        overall = df[df['Class'] == 'Overall'].copy()

        if overall.empty:
            print(f"  ⚠ No Overall metrics for {dataset_name}")
            continue

        # Get model columns
        models = [col for col in df.columns if col not in ['Class', 'Metric']]

        # Get class names for per-class metrics
        class_rows = df[df['Class'] != 'Overall']
        class_names = class_rows['Class'].unique().tolist()

        # Get training times from results.txt files
        training_times = {}
        for model in models:
            results_file = exp_folder / f'cls_{model}_focal' / 'results.txt'
            if results_file.exists():
                try:
                    with open(results_file, 'r') as f:
                        for line in f:
                            if 'Training Time:' in line:
                                # Extract time value: "Training Time: 4.1 min" -> 4.1
                                time_str = line.split('Training Time:')[1].strip().replace(' min', '')
                                training_times[model] = float(time_str)
                                break
                except:
                    training_times[model] = 0
            else:
                training_times[model] = 0

        # Transform to model-per-row format with per-class columns
        result_rows = []
        for model in models:
            model_name = model.replace('_', '-').upper() if model != 'densenet121' else 'DenseNet121'
            if 'efficientnet' in model:
                model_name = f"EfficientNet-{model.split('_')[1].upper()}"
            elif 'resnet' in model:
                model_name = f"ResNet{model.replace('resnet', '')}"

            row_data = {
                'Model': model_name,
                'Params\n(M)': model_params.get(model, 0),
                'Training\nTime (min)': training_times.get(model, 0),
                'Accuracy': 0,
                'Balanced\nAcc': 0
            }

            # Overall metrics
            for _, metric_row in overall.iterrows():
                metric = metric_row['Metric']
                value = metric_row[model]

                if isinstance(value, (int, float)):
                    if metric == 'accuracy':
                        row_data['Accuracy'] = round(value, 2)
                    elif metric == 'balanced_accuracy':
                        row_data['Balanced\nAcc'] = round(value, 2)

            # Per-class metrics (Precision, F1) - KINETIK format
            for class_name in class_names:
                class_df = df[df['Class'] == class_name]

                # Count samples for this class from support
                n_samples = 0
                for _, metric_row in class_df.iterrows():
                    if metric_row['Metric'] == 'support':
                        n_samples = int(metric_row[model])
                        break

                # Get precision and F1 for this class
                precision_val = None
                f1_val = None

                for _, metric_row in class_df.iterrows():
                    metric = metric_row['Metric']
                    value = metric_row[model]

                    if isinstance(value, (int, float)):
                        if metric == 'precision':
                            precision_val = round(value, 2)
                        elif metric == 'f1_score':
                            f1_val = round(value, 2)

                # Add columns: "{ClassName}\n(n=X)" for Precision and F1
                class_header = f"{class_name.title()}\n(n={n_samples})"
                row_data[f"{class_header}\nPrecision"] = precision_val
                row_data[f"{class_header}\nF1"] = f1_val

            result_rows.append(row_data)

        result_df = pd.DataFrame(result_rows)

        # Save individual table
        output_file = output_folder / f'{table_num}_{dataset_key}.xlsx'

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name=dataset_name, index=False)
            apply_academic_styling(writer.sheets[dataset_name])

        print(f"  ✓ Created: {output_file.name} ({len(result_df)} models)")
        table_count += 1

    print(f"  Total: {table_count} classification tables generated")

def main():
    """Main execution"""
    print("="*60)
    print("ACADEMIC JOURNAL TABLE GENERATOR")
    print("Style: KINETIK Journal Format")
    print("="*60)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent

        # AUTO-DETECT LATEST EXPERIMENT FOLDER
        results_dir = base_dir / 'results'
        experiment_folders = sorted([f for f in results_dir.glob('optA_*') if f.is_dir()], reverse=True)
        if not experiment_folders:
            raise ValueError("No experiment folders found in results/")

        experiment_folder = experiment_folders[0]  # Latest folder
        output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'
        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\n📁 Latest Experiment: {experiment_folder.name}")
        print(f"📁 Output: {output_folder}\n")

        # AUTO-DELETE OLD TABLES FIRST
        print("🗑️  Deleting old tables...")
        deleted_count = 0
        for old_file in output_folder.glob('*.xlsx'):
            old_file.unlink()
            deleted_count += 1
        for old_file in output_folder.glob('*.csv'):
            old_file.unlink()
            deleted_count += 1
        print(f"   Deleted {deleted_count} old files\n")

        # Generate tables
        generate_dataset_statistics(experiment_folder, output_folder)
        generate_detection_performance(experiment_folder, output_folder)
        generate_classification_performance(experiment_folder, output_folder)

        print("\n" + "="*60)
        print("ALL TABLES GENERATED SUCCESSFULLY!")
        print("="*60)
        print(f"\nStyle: Academic journal format (KINETIK paper)")
        print(f"Format: Table 1 (datasets) + Table 2 (detection) + Tables 3-6 (classification per dataset)")
        print(f"\nLocation: {output_folder}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
