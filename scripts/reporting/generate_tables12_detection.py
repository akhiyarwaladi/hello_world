"""
Generate Dataset Statistics and Detection Performance Tables
=============================================================

Generates Table 1 (Dataset Statistics) and Table 2 (Detection Performance)
from experiment results.

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
import argparse

# Professional color scheme
COLORS = {
    'header_bg': '3C7A8C',
    'header_text': 'FFFFFF',
    'alt_row': 'F0F4F8',
    'border': 'CBD5E0',
}

def apply_professional_styling(ws):
    """Apply professional styling to worksheet"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Header styling
    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    for row in range(2, ws.max_row + 1):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center',
                                      vertical='center')

            # Number formatting
            if isinstance(cell.value, (int, float)) and col > 1:
                cell.number_format = '0.00'

    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 50)

    # Freeze header
    ws.freeze_panes = 'A2'

def generate_dataset_statistics(experiment_folder, output_folder):
    """Generate dataset statistics table"""
    print("\nGenerating Table 1: Dataset Statistics...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'dataset_statistics_all.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    output_file = output_folder / 'Table1_Dataset_Statistics.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dataset Statistics', index=False)
        apply_professional_styling(writer.sheets['Dataset Statistics'])

    print(f"  ✓ Created: {output_file.name}")

def generate_detection_tables(experiment_folder, output_folder):
    """Generate detection performance tables"""
    print("\nGenerating Table 2: Detection Performance...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    datasets = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Consolidated file - Table 2
    output_file_all = output_folder / 'Table2_Detection_Performance.xlsx'

    with pd.ExcelWriter(output_file_all, engine='openpyxl') as writer:
        for dataset_key, dataset_name in datasets.items():
            dataset_df = df[df['Dataset'] == dataset_key].copy()
            if dataset_df.empty:
                continue

            dataset_df = dataset_df[['Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch']]
            dataset_df.columns = ['Model', 'mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)', 'Best Epoch']

            sheet_name = dataset_name[:31]
            dataset_df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_professional_styling(writer.sheets[sheet_name])

    print(f"  ✓ Created: {output_file_all.name}")

def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description='Generate Tables 1 and 2 from experiment results')
    parser.add_argument('--experiment', type=str, default='optA_20251207_233941',
                       help='Experiment folder name (default: optA_20251207_233941)')
    parser.add_argument('--output', type=str, default=None,
                       help='Output folder (default: luaran/laporan_akhir/tables)')

    args = parser.parse_args()

    print("=" * 70)
    print("DATASET & DETECTION TABLE GENERATOR")
    print("=" * 70)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent
        experiment_folder = base_dir / 'results' / args.experiment

        if not experiment_folder.exists():
            print(f"\n❌ Experiment folder not found: {experiment_folder}")
            sys.exit(1)

        if args.output:
            output_folder = Path(args.output)
        else:
            output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'

        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}")

        generate_dataset_statistics(experiment_folder, output_folder)
        generate_detection_tables(experiment_folder, output_folder)

        print("\n" + "=" * 70)
        print("TABLES 1 & 2 GENERATED SUCCESSFULLY!")
        print("=" * 70)
        print(f"\n📁 Tables saved to: {output_folder}")
        print(f"📊 Source experiment: {args.experiment}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
