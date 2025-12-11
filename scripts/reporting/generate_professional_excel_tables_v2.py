"""
Generate Professional Excel Tables for Final Report - V2
========================================================

Handles the actual format of table9 files (pivot format with Class/Metric rows).
Creates beautifully formatted Excel tables with professional styling.

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

# Professional color scheme - Eye-friendly palette
COLORS = {
    'header_bg': '3C7A8C',        # Medium teal-blue (professional, slightly darker)
    'header_text': 'FFFFFF',      # White text
    'alt_row': 'F0F4F8',          # Very light cool grey-blue
    'border': 'CBD5E0',           # Medium grey border
    'highlight': 'D6EAF8',        # Light blue for special cells (optional)
}

def apply_professional_styling(ws):
    """Apply professional styling to worksheet"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Header styling
    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    for row in range(2, ws.max_row + 1):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center',
                                      vertical='center')

            # Number formatting
            if isinstance(cell.value, (int, float)) and col > 1:
                cell.number_format = '0.00'

    # Auto-adjust column widths
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 50)

    # Freeze header
    ws.freeze_panes = 'A2'

def generate_classification_tables(experiment_folder, output_folder):
    """Generate classification performance tables from table9 files"""
    print("Generating Classification Performance Tables...")

    datasets = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    all_data = {}

    for dataset_key, dataset_name in datasets.items():
        print(f"\n  Processing {dataset_name}...")

        exp_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}'
        table_file = exp_folder / 'table9_focal_loss.csv'

        if not table_file.exists():
            print(f"    ⚠ Not found: {table_file}")
            continue

        # Read pivot table
        df = pd.read_csv(table_file)

        # Extract overall metrics
        overall = df[df['Class'] == 'Overall'].copy()

        if overall.empty:
            print(f"    ⚠ No Overall metrics found")
            continue

        # Get model columns
        models = [col for col in df.columns if col not in ['Class', 'Metric']]

        # Transform to model-per-row format
        rows = []
        for model in models:
            row_data = {'Model': model.upper().replace('_', '-')}

            for _, metric_row in overall.iterrows():
                metric = metric_row['Metric']
                value = metric_row[model]

                # Convert to percentage
                if isinstance(value, (int, float)):
                    row_data[metric] = round(value * 100, 2)

            rows.append(row_data)

        result_df = pd.DataFrame(rows)

        # Rename columns
        result_df.columns = [
            'Model' if col == 'Model' else
            f'{col.replace("_", " ").title()} (%)' for col in result_df.columns
        ]

        # Save individual dataset file with numbered naming
        table_numbers = {
            'iml_lifecycle': '3',
            'mp_idb_species': '4',
            'mp_idb_stages': '5',
            'md_2019_stages': '6'
        }
        table_num = table_numbers.get(dataset_key, 'X')
        output_file = output_folder / f'Table{table_num}_{dataset_key}.xlsx'

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, sheet_name=dataset_name, index=False)
            apply_professional_styling(writer.sheets[dataset_name])

        print(f"    ✓ Created: {output_file.name}")

        # Store for consolidated
        all_data[dataset_key] = (dataset_name, result_df)

    # Create consolidated file
    if all_data:
        print("\n  Creating consolidated file...")
        output_file_all = output_folder / 'classification_all_datasets.xlsx'

        with pd.ExcelWriter(output_file_all, engine='openpyxl') as writer:
            for dataset_key, (dataset_name, df) in all_data.items():
                sheet_name = dataset_name[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                apply_professional_styling(writer.sheets[sheet_name])

        print(f"  ✓ Created: {output_file_all.name}")

def generate_detection_tables(experiment_folder, output_folder):
    """Generate detection performance tables"""
    print("\nGenerating Detection Performance Tables...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    datasets = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Individual files
    for dataset_key, dataset_name in datasets.items():
        dataset_df = df[df['Dataset'] == dataset_key].copy()

        if dataset_df.empty:
            continue

        # Select and rename columns
        dataset_df = dataset_df[['Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch']]
        dataset_df.columns = ['Model', 'mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)', 'Best Epoch']

        # Individual files not needed for detection (only consolidated)
        # output_file = output_folder / f'detection_{dataset_key}.xlsx'

    # Consolidated file - Table 2
    output_file_all = output_folder / 'Table2_Detection_Performance.xlsx'

    with pd.ExcelWriter(output_file_all, engine='openpyxl') as writer:
        for dataset_key, dataset_name in datasets.items():
            dataset_df = df[df['Dataset'] == dataset_key].copy()
            if dataset_df.empty:
                continue

            dataset_df = dataset_df[['Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch']]
            dataset_df.columns = ['Model', 'mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)', 'Best Epoch']

            sheet_name = dataset_name[:31]
            dataset_df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_professional_styling(writer.sheets[sheet_name])

    print(f"  ✓ Created: {output_file_all.name}")

def generate_dataset_statistics(experiment_folder, output_folder):
    """Generate dataset statistics table"""
    print("\nGenerating Dataset Statistics Table...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'dataset_statistics_all.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    # Note: Duplicates fixed at root cause in main_pipeline.py (dataset_statistics_analyzer call)
    # Each experiment folder now only analyzes its own dataset

    output_file = output_folder / 'Table1_Dataset_Statistics.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dataset Statistics', index=False)
        apply_professional_styling(writer.sheets['Dataset Statistics'])

    print(f"  ✓ Created: {output_file.name}")

def main():
    """Main execution"""
    print("="*70)
    print("PROFESSIONAL EXCEL TABLE GENERATOR V2")
    print("="*70)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent
        experiment_folder = base_dir / 'results' / 'optA_20251207_233941'
        output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'
        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}\n")

        generate_dataset_statistics(experiment_folder, output_folder)
        generate_detection_tables(experiment_folder, output_folder)
        generate_classification_tables(experiment_folder, output_folder)

        print("\n" + "="*70)
        print("ALL TABLES GENERATED SUCCESSFULLY!")
        print("="*70)
        print(f"\n📁 Tables saved to: {output_folder}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
