"""
Generate Consolidated Detection Performance Table for Journal Format
====================================================================

Creates a SINGLE table with all datasets for easier comparison in publications.

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
import argparse

# Professional color scheme
COLORS = {
    'header_bg': '3C7A8C',
    'header_text': 'FFFFFF',
    'alt_row': 'F0F4F8',
    'border': 'CBD5E0',
}

def apply_professional_styling(ws):
    """Apply professional styling to worksheet"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Header styling
    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows with grouping (alternate fill per dataset)
    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    current_dataset = None
    use_fill = False

    for row in range(2, ws.max_row + 1):
        # Check if dataset changed (for grouping color)
        dataset_cell = ws.cell(row=row, column=1).value
        if dataset_cell and dataset_cell != current_dataset:
            current_dataset = dataset_cell
            use_fill = not use_fill

        fill = alt_fill if use_fill else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col <= 2 else 'center',
                                      vertical='center')

            # Number formatting
            if isinstance(cell.value, (int, float)) and col > 2:
                if col == ws.max_column:  # Best Epoch (integer)
                    cell.number_format = '0'
                else:  # Percentages
                    cell.number_format = '0.00'

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20  # Dataset
    ws.column_dimensions['B'].width = 12  # Model
    ws.column_dimensions['C'].width = 14  # mAP@50
    ws.column_dimensions['D'].width = 16  # mAP@50-95
    ws.column_dimensions['E'].width = 14  # Precision
    ws.column_dimensions['F'].width = 12  # Recall
    ws.column_dimensions['G'].width = 14  # Best Epoch

    # Freeze header
    ws.freeze_panes = 'A2'

def generate_consolidated_detection_table(experiment_folder, output_folder):
    """Generate consolidated detection table with all datasets"""
    print("\nGenerating Table 2: Consolidated Detection Performance...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return

    df = pd.read_csv(csv_file)

    # Map dataset keys to display names
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Replace dataset names
    df['Dataset'] = df['Dataset'].map(dataset_names)

    # Select and order columns
    df_out = df[['Dataset', 'Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch']].copy()

    # Rename columns with percentage indicators
    df_out.columns = ['Dataset', 'Model', 'mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)', 'Best Epoch']

    # Convert to percentages (multiply by 100)
    for col in ['mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)']:
        df_out[col] = df_out[col] * 100

    # Sort by dataset then model
    dataset_order = ['IML Lifecycle', 'MP-IDB Species', 'MP-IDB Stages', 'MD-2019 Stages']
    df_out['Dataset'] = pd.Categorical(df_out['Dataset'], categories=dataset_order, ordered=True)
    df_out = df_out.sort_values(['Dataset', 'Model'])

    # Create Excel file
    output_file = output_folder / 'Table2_Detection_Performance.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detection Performance'

    # Write headers
    headers = df_out.columns.tolist()
    ws.append(headers)

    # Write data
    for _, row in df_out.iterrows():
        ws.append(row.tolist())

    # Apply styling
    apply_professional_styling(ws)

    # Save
    wb.save(output_file)
    print(f"  ✓ Created: {output_file.name}")
    print(f"  ✓ Format: Single consolidated table with {len(df_out)} rows")
    print(f"  ✓ Datasets: 4 (IML, MP-IDB Species, MP-IDB Stages, MD-2019)")
    print(f"  ✓ Models per dataset: 3 (YOLO10, YOLO11, YOLO12)")

def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description='Generate consolidated detection table')
    parser.add_argument('--experiment', type=str, default='optA_20251207_233941',
                       help='Experiment folder name')
    parser.add_argument('--output', type=str, default=None,
                       help='Output folder (default: luaran/laporan_akhir/tables)')

    args = parser.parse_args()

    print("=" * 70)
    print("CONSOLIDATED DETECTION TABLE GENERATOR (JOURNAL FORMAT)")
    print("=" * 70)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent
        experiment_folder = base_dir / 'results' / args.experiment

        if not experiment_folder.exists():
            print(f"\n❌ Experiment folder not found: {experiment_folder}")
            sys.exit(1)

        if args.output:
            output_folder = Path(args.output)
        else:
            output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'

        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}")

        generate_consolidated_detection_table(experiment_folder, output_folder)

        print("\n" + "=" * 70)
        print("TABLE 2 GENERATED SUCCESSFULLY!")
        print("=" * 70)
        print(f"\n📁 Table saved to: {output_folder}")
        print(f"📊 Source experiment: {args.experiment}")
        print(f"📋 Format: Single consolidated table (better for journal papers)")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
