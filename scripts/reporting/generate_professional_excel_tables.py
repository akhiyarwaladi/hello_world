"""
Generate Professional Excel Tables for Final Report
===================================================

This script generates beautifully formatted Excel tables for the final research report.
All tables are sourced from the latest complete experiment folder and styled professionally.

Features:
- Auto-adjusted column widths
- Professional color schemes
- Clean borders and alignment
- Separate sheets for each dataset/metric
- Centralized data source from latest experiment

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import json
import sys

# Professional color scheme
COLORS = {
    'header_bg': 'E67E22',        # Professional Orange
    'header_text': 'FFFFFF',       # White
    'alt_row': 'FFFAE6',          # Very light orange/yellow
    'border': 'BDC3C7',           # Light grey
    'accent': '3498DB'            # Professional Blue
}

def setup_paths():
    """Setup all required paths"""
    base_dir = Path(__file__).resolve().parent.parent.parent

    # Latest experiment folder
    experiment_folder = base_dir / 'results' / 'optA_20251207_233941'

    # Output folder for tables
    output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'
    output_folder.mkdir(parents=True, exist_ok=True)

    return experiment_folder, output_folder

def apply_professional_styling(ws, start_row=1, start_col=1, end_row=None, end_col=None):
    """
    Apply professional styling to a worksheet

    Args:
        ws: Worksheet object
        start_row: Starting row for styling
        start_col: Starting column for styling
        end_row: Ending row (None = last row with data)
        end_col: Ending column (None = last column with data)
    """
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column

    # Define borders
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Header styling (first row)
    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows styling
    for row in range(start_row + 1, end_row + 1):
        # Alternate row coloring
        if (row - start_row) % 2 == 0:
            fill = PatternFill(start_color=COLORS['alt_row'],
                              end_color=COLORS['alt_row'],
                              fill_type='solid')
        else:
            fill = PatternFill(fill_type=None)

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == start_col else 'center',
                                      vertical='center')

            # Number formatting for numeric cells
            if isinstance(cell.value, (int, float)) and col > start_col:
                cell.number_format = '0.00'

    # Auto-adjust column widths
    for col in range(start_col, end_col + 1):
        column_letter = get_column_letter(col)

        # Calculate max width
        max_length = 0
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)

        # Set column width (add padding)
        adjusted_width = min(max_length + 3, 50)  # Max 50 chars
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)

def generate_dataset_statistics_excel(experiment_folder, output_folder):
    """Generate professional Excel for dataset statistics"""
    print("Generating Dataset Statistics Excel...")

    # Read the consolidated CSV
    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'dataset_statistics_all.csv'

    if not csv_file.exists():
        print(f"Warning: {csv_file} not found!")
        return

    df = pd.read_csv(csv_file)

    # Create Excel file with professional styling
    output_file = output_folder / 'dataset_statistics_all.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dataset Statistics', index=False)

        # Get worksheet and apply styling
        ws = writer.sheets['Dataset Statistics']
        apply_professional_styling(ws)

    print(f"✓ Created: {output_file.name}")

def generate_detection_performance_excel(experiment_folder, output_folder):
    """Generate professional Excel for detection performance - per dataset"""
    print("\nGenerating Detection Performance Excel (per dataset)...")

    datasets = ['iml_lifecycle', 'mp_idb_species', 'mp_idb_stages', 'md_2019_stages']
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Read consolidated CSV
    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"Warning: {csv_file} not found!")
        return

    df = pd.read_csv(csv_file)

    # Create separate Excel file for each dataset
    for dataset_key in datasets:
        dataset_df = df[df['Dataset'] == dataset_key].copy()

        if dataset_df.empty:
            continue

        # Clean up dataset name in the dataframe
        dataset_df['Dataset'] = dataset_df['Dataset'].map(dataset_names)

        # Select key columns (only those that exist)
        potential_columns = ['Dataset', 'Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch', 'Epochs']
        columns_to_keep = [col for col in potential_columns if col in dataset_df.columns]
        dataset_df = dataset_df[columns_to_keep]

        # Rename columns for clarity
        rename_map = {
            'Dataset': 'Dataset',
            'Model': 'Model',
            'mAP@50': 'mAP@50 (%)',
            'mAP@50-95': 'mAP@50-95 (%)',
            'Precision': 'Precision (%)',
            'Recall': 'Recall (%)',
            'Best Epoch': 'Best Epoch',
            'Epochs': 'Total Epochs'
        }
        dataset_df.columns = [rename_map.get(col, col) for col in dataset_df.columns]

        # Create Excel file
        output_file = output_folder / f'detection_performance_{dataset_key}.xlsx'

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            dataset_df.to_excel(writer, sheet_name=dataset_names[dataset_key], index=False)

            # Get worksheet and apply styling
            ws = writer.sheets[dataset_names[dataset_key]]
            apply_professional_styling(ws)

        print(f"✓ Created: {output_file.name}")

    # Also create a consolidated single-file version
    output_file_all = output_folder / 'detection_performance_all_datasets.xlsx'

    with pd.ExcelWriter(output_file_all, engine='openpyxl') as writer:
        for dataset_key in datasets:
            dataset_df = df[df['Dataset'] == dataset_key].copy()
            if dataset_df.empty:
                continue

            dataset_df['Dataset'] = dataset_df['Dataset'].map(dataset_names)
            potential_columns = ['Dataset', 'Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch', 'Epochs']
            columns_to_keep = [col for col in potential_columns if col in dataset_df.columns]
            dataset_df = dataset_df[columns_to_keep]

            rename_map = {
                'Dataset': 'Dataset',
                'Model': 'Model',
                'mAP@50': 'mAP@50 (%)',
                'mAP@50-95': 'mAP@50-95 (%)',
                'Precision': 'Precision (%)',
                'Recall': 'Recall (%)',
                'Best Epoch': 'Best Epoch',
                'Epochs': 'Total Epochs'
            }
            dataset_df.columns = [rename_map.get(col, col) for col in dataset_df.columns]

            sheet_name = dataset_names[dataset_key][:31]  # Excel sheet name limit
            dataset_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            apply_professional_styling(ws)

    print(f"✓ Created: {output_file_all.name} (consolidated)")

def generate_classification_performance_excel(experiment_folder, output_folder):
    """Generate professional Excel for classification performance - per dataset"""
    print("\nGenerating Classification Performance Excel (per dataset)...")

    datasets = ['iml_lifecycle', 'mp_idb_species', 'mp_idb_stages', 'md_2019_stages']
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Create separate Excel file for each dataset
    for dataset_key in datasets:
        print(f"  Processing {dataset_names[dataset_key]}...")

        # Find the experiment folder for this dataset
        exp_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}'

        if not exp_folder.exists():
            print(f"    Warning: {exp_folder} not found!")
            continue

        # Look for table9 (Focal Loss performance)
        table_files = list(exp_folder.glob('table9_*.csv'))

        if not table_files:
            print(f"    Warning: No table9 files found in {exp_folder}!")
            continue

        # Use the first table9 file found
        csv_file = table_files[0]
        df = pd.read_csv(csv_file)

        # Clean up column names
        df.columns = df.columns.str.strip()

        # Select key columns
        if 'Model' in df.columns:
            columns_to_keep = ['Model', 'Accuracy', 'Balanced_Accuracy', 'Precision_Macro', 'Recall_Macro', 'F1_Macro']

            # Check which columns actually exist
            columns_to_keep = [col for col in columns_to_keep if col in df.columns]

            dataset_df = df[columns_to_keep].copy()

            # Rename for clarity
            rename_dict = {
                'Model': 'Model',
                'Accuracy': 'Accuracy (%)',
                'Balanced_Accuracy': 'Balanced Accuracy (%)',
                'Precision_Macro': 'Precision (Macro) (%)',
                'Recall_Macro': 'Recall (Macro) (%)',
                'F1_Macro': 'F1-Score (Macro) (%)'
            }

            dataset_df.columns = [rename_dict.get(col, col) for col in dataset_df.columns]

            # Create Excel file
            output_file = output_folder / f'classification_focal_loss_{dataset_key}.xlsx'

            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                dataset_df.to_excel(writer, sheet_name=dataset_names[dataset_key], index=False)

                # Get worksheet and apply styling
                ws = writer.sheets[dataset_names[dataset_key]]
                apply_professional_styling(ws)

            print(f"    ✓ Created: {output_file.name}")

    # Create consolidated version
    print("\n  Creating consolidated classification table...")
    output_file_all = output_folder / 'classification_focal_loss_all_datasets.xlsx'

    # Collect all data first
    consolidated_data = {}

    for dataset_key in datasets:
        exp_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}'
        if not exp_folder.exists():
            continue

        table_files = list(exp_folder.glob('table9_*.csv'))
        if not table_files:
            continue

        csv_file = table_files[0]
        df = pd.read_csv(csv_file)
        df.columns = df.columns.str.strip()

        if 'Model' in df.columns:
            columns_to_keep = ['Model', 'Accuracy', 'Balanced_Accuracy', 'Precision_Macro', 'Recall_Macro', 'F1_Macro']
            columns_to_keep = [col for col in columns_to_keep if col in df.columns]
            dataset_df = df[columns_to_keep].copy()

            rename_dict = {
                'Model': 'Model',
                'Accuracy': 'Accuracy (%)',
                'Balanced_Accuracy': 'Balanced Accuracy (%)',
                'Precision_Macro': 'Precision (Macro) (%)',
                'Recall_Macro': 'Recall (Macro) (%)',
                'F1_Macro': 'F1-Score (Macro) (%)'
            }

            dataset_df.columns = [rename_dict.get(col, col) for col in dataset_df.columns]
            consolidated_data[dataset_key] = dataset_df

    # Only create consolidated file if we have data
    if consolidated_data:
        with pd.ExcelWriter(output_file_all, engine='openpyxl') as writer:
            for dataset_key, dataset_df in consolidated_data.items():
                sheet_name = dataset_names[dataset_key][:31]
                dataset_df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]
                apply_professional_styling(ws)

        print(f"  ✓ Created: {output_file_all.name} (consolidated)")
    else:
        print(f"  ⚠ Warning: No data found for consolidated table!")

def generate_training_time_comparison_excel(experiment_folder, output_folder):
    """Generate professional Excel for training time comparison"""
    print("\nGenerating Training Time Comparison Excel...")

    datasets = ['iml_lifecycle', 'mp_idb_species', 'mp_idb_stages', 'md_2019_stages']
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }

    # Collect training time data from each experiment
    training_data = []

    for dataset_key in datasets:
        exp_folder = experiment_folder / 'experiments' / f'experiment_{dataset_key}'

        if not exp_folder.exists():
            continue

        # Look for classification model folders
        cls_folders = list(exp_folder.glob('cls_*_focal'))

        for cls_folder in cls_folders:
            # Extract model name from folder
            model_name = cls_folder.name.replace('cls_', '').replace('_focal', '')

            # Look for training log or summary
            summary_file = cls_folder / 'training_summary.json'

            if summary_file.exists():
                with open(summary_file, 'r') as f:
                    summary = json.load(f)

                training_time = summary.get('total_training_time', 0)

                # Convert seconds to hours
                training_time_hours = training_time / 3600

                training_data.append({
                    'Dataset': dataset_names[dataset_key],
                    'Model': model_name,
                    'Training Time (hours)': round(training_time_hours, 2),
                    'Epochs': summary.get('epochs', 75)
                })

    if not training_data:
        print("  Warning: No training time data found!")
        return

    df = pd.DataFrame(training_data)

    # Create Excel file
    output_file = output_folder / 'training_time_comparison.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Create a pivot table style view
        pivot_df = df.pivot(index='Model', columns='Dataset', values='Training Time (hours)')
        pivot_df.to_excel(writer, sheet_name='Training Time Comparison')

        ws = writer.sheets['Training Time Comparison']
        apply_professional_styling(ws, start_col=1)

        # Also create a detailed sheet
        df.to_excel(writer, sheet_name='Detailed Data', index=False)
        ws_detail = writer.sheets['Detailed Data']
        apply_professional_styling(ws_detail)

    print(f"✓ Created: {output_file.name}")

def main():
    """Main function to generate all professional Excel tables"""
    print("="*70)
    print("GENERATING PROFESSIONAL EXCEL TABLES FOR FINAL REPORT")
    print("="*70)

    try:
        # Setup paths
        experiment_folder, output_folder = setup_paths()

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}\n")

        # Generate all tables
        generate_dataset_statistics_excel(experiment_folder, output_folder)
        generate_detection_performance_excel(experiment_folder, output_folder)
        generate_classification_performance_excel(experiment_folder, output_folder)
        generate_training_time_comparison_excel(experiment_folder, output_folder)

        print("\n" + "="*70)
        print("ALL PROFESSIONAL EXCEL TABLES GENERATED SUCCESSFULLY!")
        print("="*70)
        print(f"\nTables location: {output_folder}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
