"""
Generate All Tables (1-6) - Centralized Table Generator
========================================================

Master script that generates all 6 tables in correct format:
- Table 1: Dataset Statistics
- Table 2: Detection Performance (Consolidated)
- Tables 3-6: Classification Performance (KINETIK format)

Author: Research Team
Date: December 2025
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
import argparse
import re
import shutil

# Professional color scheme
COLORS = {
    'header_bg': '3C7A8C',
    'header_text': 'FFFFFF',
    'alt_row': 'F0F4F8',
    'border': 'CBD5E0',
}

def clean_output_folder(output_folder):
    """Delete all existing .xlsx files in output folder"""
    print("\n🗑️  Cleaning existing tables...")
    deleted_count = 0
    for file in output_folder.glob('*.xlsx'):
        file.unlink()
        deleted_count += 1
        print(f"  ✓ Deleted: {file.name}")

    if deleted_count == 0:
        print("  ℹ️  No existing tables to delete")
    else:
        print(f"  ✓ Deleted {deleted_count} existing tables")

# ============================================================================
# TABLE 1: Dataset Statistics
# ============================================================================

def apply_statistics_styling(ws):
    """Apply professional styling for statistics table"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    for row in range(2, ws.max_row + 1):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'center',
                                      vertical='center')

            # Smart number formatting: integers without decimals, decimals with 2 places
            if isinstance(cell.value, (int, float)) and col > 1:
                if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value.is_integer()):
                    cell.number_format = '0'  # No decimals for whole numbers
                else:
                    cell.number_format = '0.00'  # 2 decimals for fractional numbers

    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 50)

    ws.freeze_panes = 'A2'

def generate_table1(experiment_folder, output_folder):
    """Generate Table 1: Dataset Statistics"""
    print("\n📊 Generating Table 1: Dataset Statistics...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'dataset_statistics_all.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return False

    df = pd.read_csv(csv_file)

    # Remove duplicates
    original_rows = len(df)
    df = df.drop_duplicates(subset=['Dataset'], keep='first')
    if len(df) < original_rows:
        print(f"  ⚠ Removed {original_rows - len(df)} duplicate rows ({original_rows} → {len(df)})")

    # Standardize dataset names
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }
    df['Dataset'] = df['Dataset'].map(dataset_names)

    output_file = output_folder / 'Table1_Dataset_Statistics.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dataset Statistics', index=False)
        apply_statistics_styling(writer.sheets['Dataset Statistics'])

    print(f"  ✓ Created: {output_file.name}")
    return True

# ============================================================================
# TABLE 2: Detection Performance (Consolidated)
# ============================================================================

def apply_detection_styling(ws):
    """Apply professional styling for detection table"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    current_dataset = None
    use_fill = False

    for row in range(2, ws.max_row + 1):
        dataset_cell = ws.cell(row=row, column=1).value
        if dataset_cell and dataset_cell != current_dataset:
            current_dataset = dataset_cell
            use_fill = not use_fill

        fill = alt_fill if use_fill else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col <= 2 else 'center',
                                      vertical='center')

            if isinstance(cell.value, (int, float)) and col > 2:
                if col == ws.max_column:  # Best Epoch (always integer)
                    cell.number_format = '0'
                else:  # Percentages (smart formatting)
                    if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value.is_integer()):
                        cell.number_format = '0'  # No decimals for whole numbers
                    else:
                        cell.number_format = '0.00'  # 2 decimals for fractional numbers

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14

    ws.freeze_panes = 'A2'

def generate_table2(experiment_folder, output_folder):
    """Generate Table 2: Detection Performance (Consolidated)"""
    print("\n📊 Generating Table 2: Detection Performance...")

    csv_file = experiment_folder / 'consolidated_analysis' / 'cross_dataset_comparison' / 'detection_performance_all_datasets.csv'

    if not csv_file.exists():
        print(f"  ⚠ Not found: {csv_file}")
        return False

    df = pd.read_csv(csv_file)

    # Standardize dataset names
    dataset_names = {
        'iml_lifecycle': 'IML Lifecycle',
        'mp_idb_species': 'MP-IDB Species',
        'mp_idb_stages': 'MP-IDB Stages',
        'md_2019_stages': 'MD-2019 Stages'
    }
    df['Dataset'] = df['Dataset'].map(dataset_names)

    # Select and order columns
    df_out = df[['Dataset', 'Model', 'mAP@50', 'mAP@50-95', 'Precision', 'Recall', 'Best Epoch']].copy()
    df_out.columns = ['Dataset', 'Model', 'mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)', 'Best Epoch']

    # Convert to percentages
    for col in ['mAP@50 (%)', 'mAP@50-95 (%)', 'Precision (%)', 'Recall (%)']:
        df_out[col] = df_out[col] * 100

    # Sort by dataset then model
    dataset_order = ['IML Lifecycle', 'MP-IDB Species', 'MP-IDB Stages', 'MD-2019 Stages']
    df_out['Dataset'] = pd.Categorical(df_out['Dataset'], categories=dataset_order, ordered=True)
    df_out = df_out.sort_values(['Dataset', 'Model'])

    output_file = output_folder / 'Table2_Detection_Performance.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detection Performance'

    headers = df_out.columns.tolist()
    ws.append(headers)

    for _, row in df_out.iterrows():
        ws.append(row.tolist())

    apply_detection_styling(ws)
    wb.save(output_file)

    print(f"  ✓ Created: {output_file.name}")
    return True

# ============================================================================
# TABLES 3-6: Classification Performance (KINETIK Format)
# ============================================================================

def extract_training_time(results_file):
    """Extract training time from results.txt file"""
    try:
        with open(results_file, 'r') as f:
            content = f.read()
            match = re.search(r'Training Time:\s*([\d.]+)\s*min', content)
            if match:
                return float(match.group(1))
    except Exception as e:
        print(f"  ⚠ Could not read training time from {results_file}: {e}")
    return None

def read_dataset_data(experiment_folder, dataset_key):
    """Read classification data for a specific dataset"""
    exp_path = experiment_folder / 'experiments' / f'experiment_{dataset_key}'

    table9_path = exp_path / 'table9_focal_loss.csv'
    if not table9_path.exists():
        print(f"  ⚠ table9 not found: {table9_path}")
        return None

    df = pd.read_csv(table9_path)

    overall = df[df['Class'] == 'Overall']
    if overall.empty:
        print(f"  ⚠ No overall metrics found")
        return None

    models = [col for col in df.columns if col not in ['Class', 'Metric']]
    classes = df[df['Class'] != 'Overall']['Class'].unique().tolist()

    models_data = []
    for model in models:
        accuracy = overall[overall['Metric'] == 'accuracy'][model].values[0]
        balanced_acc = overall[overall['Metric'] == 'balanced_accuracy'][model].values[0]

        results_path = exp_path / f'cls_{model}_focal' / 'results.txt'
        training_time = extract_training_time(results_path)

        per_class_metrics = []
        for cls in classes:
            cls_data = df[df['Class'] == cls]
            precision = cls_data[cls_data['Metric'] == 'precision'][model].values[0]
            f1 = cls_data[cls_data['Metric'] == 'f1_score'][model].values[0]
            support = int(cls_data[cls_data['Metric'] == 'support'][model].values[0])
            per_class_metrics.append({
                'class': cls,
                'precision': precision * 100,
                'f1': f1 * 100,
                'support': support
            })

        param_map = {
            'densenet121': 8.0,
            'efficientnet_b0': 5.3,
            'efficientnet_b1': 7.8,
            'efficientnet_b2': 9.1,
            'resnet50': 25.6,
            'resnet101': 44.5
        }

        models_data.append({
            'model': model,
            'params': param_map.get(model, 0),
            'training_time': training_time,
            'accuracy': accuracy * 100,
            'balanced_acc': balanced_acc * 100,
            'per_class_metrics': per_class_metrics
        })

    return {
        'models_data': models_data,
        'classes': classes
    }

def apply_classification_styling(ws, num_classes):
    """Apply professional styling to classification worksheet"""
    thin_border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    header_fill = PatternFill(start_color=COLORS['header_bg'],
                             end_color=COLORS['header_bg'],
                             fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)

    alt_fill = PatternFill(start_color=COLORS['alt_row'],
                          end_color=COLORS['alt_row'],
                          fill_type='solid')

    # Style header rows (rows 1-2)
    for row in [1, 2]:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Style data rows
    for row in range(3, ws.max_row + 1):
        fill = alt_fill if row % 2 == 0 else PatternFill(fill_type=None)

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border

            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            if isinstance(cell.value, (int, float)) and col > 1:
                if col in [2, 3]:  # Parameters and Training Time
                    cell.number_format = '0.0'
                else:  # All metrics (smart formatting)
                    if isinstance(cell.value, int) or (isinstance(cell.value, float) and cell.value.is_integer()):
                        cell.number_format = '0'  # No decimals for whole numbers
                    else:
                        cell.number_format = '0.00'  # 2 decimals for fractional numbers

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    for i in range(num_classes * 2):
        col_letter = get_column_letter(6 + i)
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = 'A3'

def create_classification_table(output_path, sheet_name, dataset_data):
    """Create classification table with 2-level headers"""
    models_data = dataset_data['models_data']
    classes = dataset_data['classes']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Header row 1
    row1 = ['Model', 'Parameters (M)', 'Training Time (min)', 'Accuracy', 'Balanced Acc']
    ws.append(row1 + [''] * (len(classes) * 2))

    # Merge first 5 columns vertically (row 1 to row 2)
    for col in range(1, 6):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        cell = ws.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Merge cells for class names (horizontally across precision and F1)
    start_col = 6
    for i, cls_info in enumerate([m for m in models_data[0]['per_class_metrics']]):
        cls = cls_info['class']
        support = cls_info['support']
        col = start_col + i * 2
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws.cell(row=1, column=col)
        cell.value = f"{cls.replace('_', ' ').title()} (n={support})"
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Header row 2 (only for per-class metrics)
    row2 = [''] * 5
    for _ in classes:
        row2.extend(['Precision', 'F1'])
    ws.append(row2)

    # Data rows
    for model_data in models_data:
        row = [
            model_data['model'].replace('_', '-').upper(),
            model_data['params'],
            model_data['training_time'] if model_data['training_time'] else '',
            model_data['accuracy'],
            model_data['balanced_acc']
        ]

        for cls_metric in model_data['per_class_metrics']:
            row.extend([cls_metric['precision'], cls_metric['f1']])

        ws.append(row)

    apply_classification_styling(ws, num_classes=len(classes))
    wb.save(output_path)

def generate_tables3to6(experiment_folder, output_folder):
    """Generate Tables 3-6: Classification Performance"""
    print("\n📊 Generating Tables 3-6: Classification Performance...")

    datasets = {
        'iml_lifecycle': ('IML Lifecycle', 'Table3'),
        'mp_idb_species': ('MP-IDB Species', 'Table4'),
        'mp_idb_stages': ('MP-IDB Stages', 'Table5'),
        'md_2019_stages': ('MD-2019 Stages', 'Table6')
    }

    success_count = 0
    for dataset_key, (dataset_name, table_name) in datasets.items():
        print(f"  Processing {dataset_name}...")

        dataset_data = read_dataset_data(experiment_folder, dataset_key)
        if dataset_data is None:
            continue

        output_path = output_folder / f'{table_name}_{dataset_key}.xlsx'
        create_classification_table(output_path, dataset_name, dataset_data)
        print(f"  ✓ Created: {output_path.name}")
        success_count += 1

    return success_count == 4

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description='Generate all tables (1-6) from experiment results')
    parser.add_argument('--experiment', type=str, default='optA_20251207_233941',
                       help='Experiment folder name (default: optA_20251207_233941)')
    parser.add_argument('--output', type=str, default=None,
                       help='Output folder (default: luaran/laporan_akhir/tables)')

    args = parser.parse_args()

    print("=" * 70)
    print("ALL TABLES GENERATOR (1-6) - CENTRALIZED")
    print("=" * 70)

    try:
        base_dir = Path(__file__).resolve().parent.parent.parent
        experiment_folder = base_dir / 'results' / args.experiment

        if not experiment_folder.exists():
            print(f"\n❌ Experiment folder not found: {experiment_folder}")
            sys.exit(1)

        if args.output:
            output_folder = Path(args.output)
        else:
            output_folder = base_dir / 'luaran' / 'laporan_akhir' / 'tables'

        output_folder.mkdir(parents=True, exist_ok=True)

        print(f"\nSource: {experiment_folder}")
        print(f"Output: {output_folder}")

        # Clean existing tables first
        clean_output_folder(output_folder)

        # Generate all tables
        table1_ok = generate_table1(experiment_folder, output_folder)
        table2_ok = generate_table2(experiment_folder, output_folder)
        tables36_ok = generate_tables3to6(experiment_folder, output_folder)

        print("\n" + "=" * 70)
        if table1_ok and table2_ok and tables36_ok:
            print("✅ ALL 6 TABLES GENERATED SUCCESSFULLY!")
        else:
            print("⚠️  SOME TABLES FAILED TO GENERATE")
        print("=" * 70)
        print(f"\n📁 Tables saved to: {output_folder}")
        print(f"📊 Source experiment: {args.experiment}")
        print(f"📋 Generated: Table 1-6 (Dataset Stats, Detection, Classification)")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
